"""ui/main_window.py"""

import sys
from pathlib import Path

import argparse, datetime, json, logging, traceback
from typing import Any, Callable, Dict, List, Optional, Tuple
from PyQt6.QtCore import QEvent, QObject, QProcess, QSize, QThread, QTimer, QUrl, Qt, pyqtSignal
from PyQt6.QtGui import QAction, QColor, QDesktopServices, QFont, QIcon, QPainter, QPixmap
from PyQt6.QtWidgets import QApplication, QCheckBox, QDialog, QFileDialog, QFrame, QHBoxLayout, QLabel, QLineEdit, QListWidget, QListWidgetItem, QMainWindow, QMenu, QMessageBox, QProgressBar, QPushButton, QStackedWidget, QSystemTrayIcon, QToolButton, QVBoxLayout, QWidget
import analytics
import tracker
from tracker import CONFIG_FILE, DEFAULT_CONFIG, discover_mods, load_config, run_init, run_poll, save_config
from storage import Storage
import transport
from ui.theme import set_theme


def _cli_config_path() -> str:
    for i, arg in enumerate(sys.argv):
        if arg == "--config" and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return CONFIG_FILE


set_theme(str(load_config(_cli_config_path()).get("ui", {}).get("theme", "dark")))

from ui.theme import ACCENT, ERROR, GRAY, QSS, build_qss, current_theme  # noqa: E402
from ui.icons import _icon  # noqa: E402
from ui.widgets import relative_time  # noqa: E402

logger = logging.getLogger("tracker.gui")
VERSION = "2.1.0"

from ui.pages import (  # noqa: E402
    AchievementsPage,
    AnalyticsPage,
    ChartsPage,
    CommentsPage,
    ComparePage,
    DashboardPage,
    EventsPage,
    HistoryPage,
    InsightsPage,
    LogPage,
    ModsPage,
    SearchResultsPage,
    SettingsPage,
)
from ui.pages.settings import is_profile_configured, reset_app_data  # noqa: E402

def tray_icon_path() -> str:
    """Return a cached .png for the tray icon, drawing it if missing."""
    path = Path(__file__).resolve().parent / "tray_icon.png"
    if path.exists():
        return str(path)
    pixmap = QPixmap(64, 64)
    pixmap.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    painter.setBrush(QColor(ACCENT))
    painter.setPen(Qt.PenStyle.NoPen)
    painter.drawRoundedRect(4, 4, 56, 56, 14, 14)
    painter.setPen(QColor("#FFFFFF"))
    font = QFont("Segoe UI", 22, QFont.Weight.Bold)
    painter.setFont(font)
    painter.drawText(pixmap.rect(), Qt.AlignmentFlag.AlignCenter, "M")
    painter.end()
    pixmap.save(str(path))
    return str(path)



class TrackerWorker(QThread):
    done = pyqtSignal(str)
    failed = pyqtSignal(str)

    def __init__(self, config: Dict[str, Any], fn: Callable, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._config = config
        self._fn = fn

    def run(self) -> None:  # noqa: D102
        try:
            storage = Storage(self._config["paths"]["db"])
            try:
                message = self._fn(storage, self._config) or ""
                if not isinstance(message, str):
                    message = str(message)
                self.done.emit(message)
            finally:
                storage.close()
        except Exception as exc:  # noqa: BLE001
            traceback.print_exc()
            self.failed.emit(str(exc))



def run_discover(storage: Storage, config: Dict[str, Any]) -> Any:
    """Rescan the member profile for mods/addons/files (runs on the worker)."""
    from tracker import discover_mods

    found = discover_mods(storage, config)
    return f"Found {len(found)} item(s)"



class LogBridge(QObject):
    line = pyqtSignal(str)



class GuiLogHandler(logging.Handler):
    def __init__(self, bridge: LogBridge) -> None:
        super().__init__()
        self._bridge = bridge

    def emit(self, record: logging.LogRecord) -> None:  # noqa: D102
        try:
            message = self.format(record)
            self._bridge.line.emit(message)
        except Exception:  # noqa: BLE001
            pass



def show_window(window: "TrackerWindow") -> None:
    """Show the main window, maximized (keeps the native title bar) unless opted out."""
    if window.config.get("ui", {}).get("fullscreen", True):
        window.showMaximized()
    else:
        window.show()



class FirstRunDialog(QDialog):
    """Asks for the ModDB username on the very first run."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Welcome to ModDB Tracker")
        self.setModal(True)
        self.setMinimumWidth(480)

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 24, 24, 24)
        root.setSpacing(12)

        title = QLabel("Welcome to ModDB Tracker")
        title.setStyleSheet("font-size: 18px; font-weight: 800;")
        root.addWidget(title)

        hint = QLabel(
            "Enter your ModDB username to track your published mods, addons and files.\n"
            "The app will discover your content and start polling automatically.\n"
            "You can change this later in the Configuration page."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #64748B; font-size: 12px;")
        root.addWidget(hint)

        self.username = QLineEdit()
        self.username.setPlaceholderText("Your ModDB username, e.g. rr5")
        self.username.setFocus()
        root.addWidget(self.username)

        buttons = QHBoxLayout()
        skip_btn = QPushButton("Skip")
        skip_btn.clicked.connect(self.reject)
        self.ok_btn = QPushButton("Continue")
        self.ok_btn.setObjectName("Primary")
        self.ok_btn.setDefault(True)
        self.ok_btn.clicked.connect(self._accept)
        buttons.addWidget(skip_btn)
        buttons.addStretch(1)
        buttons.addWidget(self.ok_btn)
        root.addLayout(buttons)

    def _accept(self) -> None:
        if not self.username.text().strip():
            QMessageBox.warning(self, "Username required", "Enter your ModDB username to continue.")
            return
        self.accept()

    def profile_url(self) -> str:
        name = self.username.text().strip().strip("/").split("/")[-1]
        return f"https://www.moddb.com/members/{name}"



class TrackerWindow(QMainWindow):
    NAV = [
        ("Dashboard", "home"),
        ("My Mods", "package"),
        ("Analytics", "chart"),
        ("Charts", "grid"),
        ("Compare", "scale"),
        ("Insights", "bulb"),
        ("Achievements", "trophy"),
        ("History", "trend"),
        ("Comments", "chat"),
        ("Notifications", "bell"),
        ("Configuration", "gear"),
        ("Log", "file"),
    ]

    def __init__(self, config_path: str = CONFIG_FILE) -> None:
        super().__init__()
        self.config_path = config_path
        self.config = load_config(config_path)
        transport.patch_moddb()  # ensure moddb's internal get_page is Cloudflare-proof
        self.storage = Storage(self.config["paths"]["db"])
        self.worker: Optional[TrackerWorker] = None
        self._stopping = False
        self._tray_hint_shown = False
        self.tray: Optional[QSystemTrayIcon] = None

        self.setWindowTitle("ModDB Tracker")
        self.resize(1280, 820)
        self.setMinimumSize(1080, 680)

        self._search_override = False
        self._saved_page = 0
        self._build_ui()
        self._wire_logging()
        self._build_tray()

        self.auto_timer = QTimer(self)
        self.auto_timer.timeout.connect(self._poll_now)
        self.auto_checkbox = QCheckBox("Auto-poll")
        self.auto_checkbox.toggled.connect(self._toggle_auto_poll)
        self.statusBar().addWidget(self.auto_checkbox)
        self.busy_bar = QProgressBar()
        self.busy_bar.setRange(0, 0)
        self.busy_bar.setFixedWidth(120)
        self.busy_bar.setVisible(False)
        self.statusBar().addWidget(self.busy_bar)

        self.status_label = QLabel("")
        self.statusBar().addPermanentWidget(self.status_label)

        self._db_label = QLabel("")
        self.statusBar().addPermanentWidget(self._db_label)

        self.refresh_all()
        self.set_status("Ready. Click \u201cPoll now\u201d to fetch data.")

    # ---- ui construction ------------------------------------------------
    def _build_ui(self) -> None:
        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_top_bar())

        body = QHBoxLayout()
        body.setContentsMargins(14, 14, 14, 12)
        body.setSpacing(14)
        body.addWidget(self._build_sidebar())

        self.stack = QStackedWidget()
        self.dashboard = DashboardPage(self.config, self.config_path)
        self.mods = ModsPage(self.config, self.config_path)
        self.analytics = AnalyticsPage(self.config, self.config_path)
        self.charts = ChartsPage(self.config, self.config_path)
        self.compare = ComparePage(self.config, self.config_path)
        self.insights = InsightsPage()
        self.achievements = AchievementsPage()
        self.history = HistoryPage()
        self.comments = CommentsPage()
        self.events = EventsPage()
        self.log_page = LogPage()
        self.settings = SettingsPage(self.config)
        self.search_page = SearchResultsPage()
        for page in (self.dashboard, self.mods, self.analytics, self.charts, self.compare, self.insights, self.achievements, self.history, self.comments, self.events, self.settings, self.log_page):
            self.stack.addWidget(page)
        self.stack.addWidget(self.search_page)
        self.search_page.open_url.connect(self._open_url)
        self.settings.saved.connect(self._on_settings_saved)
        self.settings.file_reload.connect(self._on_settings_saved)
        self.settings.reset_requested.connect(self._reset_app_data)
        self.mods.open_requested.connect(self._open_url)
        self.mods.refresh_requested.connect(self._refresh_mod)
        self.mods.favorite_toggled.connect(self._toggle_favorite)
        self.mods.remove_requested.connect(self._remove_mod)
        self.mods.export_requested.connect(self._export_json)
        self.dashboard.view_insights.connect(self._open_insights)
        self.charts.regen_requested.connect(self._regenerate_charts)
        self.events.open_url.connect(self._open_url)
        self.events.events_read.connect(self._update_badge)
        self.history.backfill_requested.connect(self._backfill_mod)
        self.history.backfill_all_requested.connect(self._backfill_all)
        body.addWidget(self.stack, 1)
        root.addLayout(body, 1)

        self.setCentralWidget(central)
        self.sidebar.setCurrentRow(0)

    def _build_top_bar(self) -> QWidget:
        bar = QFrame()
        bar.setObjectName("TopBar")
        h = QHBoxLayout(bar)
        h.setContentsMargins(18, 10, 18, 10)
        h.setSpacing(12)

        logo = QLabel()
        logo.setPixmap(QIcon(tray_icon_path()).pixmap(30, 30))
        logo.setFixedSize(30, 30)
        title = QLabel("ModDB Tracker")
        title.setStyleSheet("font-size: 16px; font-weight: 800; letter-spacing: 0.5px;")
        version = QLabel(f"v{VERSION}")
        version.setStyleSheet(
            f"color: {ACCENT}; font-size: 10px; font-weight: 700; background: {ACCENT}22;"
            "padding: 2px 7px; border-radius: 8px;"
        )
        brand = QHBoxLayout()
        brand.setSpacing(8)
        brand.addWidget(logo)
        brand.addWidget(title)
        brand.addWidget(version, 0, Qt.AlignmentFlag.AlignVCenter)
        h.addLayout(brand)

        h.addStretch(1)

        self.search_box = QLineEdit()
        self.search_box.setObjectName("Search")
        self.search_box.setPlaceholderText("Search mods, comments, events\u2026")
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setFixedWidth(240)
        self.search_box.textChanged.connect(self._search_changed)
        h.addWidget(self.search_box)

        self.sync_label = QLabel("Not synced")
        self.sync_label.setStyleSheet(f"color: {GRAY}; font-size: 12px;")
        h.addWidget(self.sync_label)

        self.bell_btn = QToolButton()
        self.bell_btn.setIcon(QIcon(_icon("bell", "#94A3B8", 20)))
        self.bell_btn.setToolTip("Notifications")
        self.bell_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.bell_btn.clicked.connect(self._show_notifications)
        h.addWidget(self.bell_btn)
        self.badge = QLabel("")
        self.badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.badge.setMinimumWidth(18)
        self.badge.setStyleSheet(
            f"background: {ERROR}; color: #FFFFFF; font-size: 10px; font-weight: 700;"
            "border-radius: 9px; padding: 0 5px;"
        )
        self.badge.setVisible(False)
        h.addWidget(self.badge)

        self.export_btn = QToolButton()
        self.export_btn.setText("Export \u25be")
        self.export_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.export_btn.setToolTip("Export reports")
        export_menu = QMenu(self)
        export_menu.addAction("CSV\u2026", self._export_csv)
        export_menu.addAction("Excel (.xlsx)\u2026", self._export_xlsx)
        export_menu.addAction("PDF report\u2026", self._export_pdf)
        export_menu.addSeparator()
        export_menu.addAction("JSON\u2026", self._export_json)
        self.export_btn.setMenu(export_menu)
        h.addWidget(self.export_btn)

        self.poll_btn = QPushButton("Poll now")
        self.poll_btn.setObjectName("Primary")
        self.poll_btn.clicked.connect(self._poll_now)
        self.rescan_btn = QPushButton("Rescan")
        self.rescan_btn.clicked.connect(self._rescan)
        self.charts_btn = QPushButton("Charts")
        self.charts_btn.clicked.connect(self._regenerate_charts)
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.clicked.connect(self.refresh_all)
        h.addWidget(self.poll_btn)
        h.addWidget(self.rescan_btn)
        h.addWidget(self.charts_btn)
        h.addWidget(self.refresh_btn)
        return bar

    def _build_sidebar(self) -> QWidget:
        self.sidebar = QListWidget()
        self.sidebar.setObjectName("Sidebar")
        self.sidebar.setFixedWidth(200)
        for name, icon_name in self.NAV:
            item = QListWidgetItem(name)
            item.setIcon(QIcon(_icon(icon_name, "#94A3B8", 20)))
            item.setSizeHint(QSize(0, 38))
            self.sidebar.addItem(item)
        self.sidebar.currentRowChanged.connect(self._nav_changed)
        return self.sidebar

    def _nav_changed(self, row: int) -> None:
        if self._search_override:
            self._search_override = False
            self.search_box.blockSignals(True)
            self.search_box.clear()
            self.search_box.blockSignals(False)
        self.stack.setCurrentIndex(row)
        if self.stack.currentWidget() is self.events:
            self.events.mark_all_seen()

    # ---- system tray ----------------------------------------------------
    def _build_tray(self) -> None:
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return
        self.tray = QSystemTrayIcon(QIcon(tray_icon_path()), self)
        self.tray.setToolTip("ModDB Tracker")

        menu = QMenu()
        self.tray_menu_show = QAction("Show / Hide", self)
        self.tray_menu_show.triggered.connect(self._toggle_window)
        self.tray_menu_poll = QAction("Poll now", self)
        self.tray_menu_poll.triggered.connect(self._poll_now)
        self.tray_menu_rescan = QAction("Rescan profile", self)
        self.tray_menu_rescan.triggered.connect(self._rescan)
        self.tray_menu_quit = QAction("Quit", self)
        self.tray_menu_quit.triggered.connect(self.close)
        menu.addAction(self.tray_menu_show)
        menu.addAction(self.tray_menu_poll)
        menu.addAction(self.tray_menu_rescan)
        menu.addSeparator()
        menu.addAction(self.tray_menu_quit)
        self.tray.setContextMenu(menu)
        self.tray.activated.connect(self._tray_activated)
        self.tray.show()

    def _tray_activated(self, reason) -> None:
        if reason == QSystemTrayIcon.ActivationReason.Trigger:
            self._toggle_window()

    def _toggle_window(self) -> None:
        if self.isVisible():
            self.hide()
        else:
            self.show()
            self.showNormal()
            self.raise_()
            self.activateWindow()

    def hide_to_tray(self, first_time: bool = False) -> None:
        if self.tray is None:
            return
        self.hide()
        if first_time or not self._tray_hint_shown:
            self._tray_hint_shown = True
            self.tray.showMessage(
                "ModDB Tracker",
                "Still running in the background. Polling and notifications stay active.",
                QSystemTrayIcon.MessageIcon.Information,
                3000,
            )

    def changeEvent(self, event) -> None:  # noqa: N802
        if event.type() == QEvent.Type.WindowStateChange and self.isMinimized():
            if self.tray is not None and self.config.get("tray", {}).get("minimize_to_tray", True):
                event.ignore()
                self.hide_to_tray()
                return
        super().changeEvent(event)

    def closeEvent(self, event) -> None:  # noqa: N802
        self._stopping = True
        self.auto_timer.stop()
        if self.worker is not None and self.worker.isRunning():
            self.worker.wait(5000)
        self.storage.close()
        if self.tray is not None:
            self.tray.hide()
        event.accept()

    # ---- logging --------------------------------------------------------
    def _wire_logging(self) -> None:
        self._bridge = LogBridge(self)
        self._bridge.line.connect(self.log_page.appendPlainText)
        handler = GuiLogHandler(self._bridge)
        handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s"))
        logging.getLogger().addHandler(handler)

    # ---- data -----------------------------------------------------------
    def refresh_all(self) -> None:
        self.dashboard.refresh(self.storage)
        self.mods.refresh(self.storage)
        self.analytics.refresh(self.storage)
        self.charts.refresh(self.storage)
        self.compare.refresh(self.storage)
        self.insights.refresh(self.storage)
        self.achievements.refresh(self.storage)
        self.history.refresh(self.storage)
        self.comments.refresh(self.storage)
        self.events.refresh(self.storage)
        self.settings.reload()
        self._update_badge()

        member = self.storage.meta_get("member_name") or ""
        last = self.storage.meta_get("last_poll") or ""
        bits = [f"Member: {member}"] if member else ["Member: not set"]
        bits.append(f"Last poll: {last}" if last else "Never polled")
        self.status_label.setText("   |   ".join(bits))
        self.sync_label.setText(f"Synced {relative_time(last)}" if last else "Not synced")
        self._update_tray_tooltip(member, last)

        try:
            db_path = Path(self.storage.db_path) if hasattr(self.storage, "db_path") else Path(self.config["paths"]["db"])
            size_mb = db_path.stat().st_size / (1024 * 1024) if db_path.exists() else 0
            self._db_label.setText(f"DB {size_mb:.1f} MB")
        except Exception:  # noqa: BLE001
            pass

    def _on_settings_saved(self) -> None:
        self.config = load_config(self.config_path)
        self.storage = Storage(self.config["paths"]["db"])
        self.dashboard._config = self.config
        self.dashboard._load_prefs()
        self.dashboard._apply_layout()
        self.mods._config = self.config
        self.analytics._config = self.config
        self.charts._config = self.config
        self.compare._config = self.config
        self._restart_auto_timer()
        self.refresh_all()

    def _open_url(self, url: str) -> None:
        if url:
            QDesktopServices.openUrl(QUrl(url))

    # ---- actions --------------------------------------------------------
    def _set_busy(self, busy: bool, label: str = "") -> None:
        self.poll_btn.setEnabled(not busy)
        self.rescan_btn.setEnabled(not busy)
        self.charts_btn.setEnabled(not busy)
        self.refresh_btn.setEnabled(not busy)
        self.export_btn.setEnabled(not busy)
        self.charts.regen_btn.setEnabled(not busy)
        self.busy_bar.setVisible(busy)
        if busy:
            self.set_status(label or "Working...")

    def _update_badge(self) -> None:
        count = self.storage.count_unseen()
        self.badge.setText(f"{count}" if count < 100 else "99+")
        self.badge.setVisible(count > 0)
        self.bell_btn.setToolTip(f"{count} new notification(s)" if count else "Notifications")

    def _show_notifications(self) -> None:
        row = next(i for i, (name, _) in enumerate(self.NAV) if name == "Notifications")
        self.sidebar.setCurrentRow(row)
        self.stack.setCurrentIndex(row)
        self.events.mark_all_seen()

    def _open_insights(self) -> None:
        row = next(i for i, (name, _) in enumerate(self.NAV) if name == "Insights")
        self.sidebar.setCurrentRow(row)
        self.stack.setCurrentIndex(row)

    def set_status(self, text: str) -> None:
        self.status_label.setText(text)
        if self.tray is not None:
            self.tray.setToolTip(f"ModDB Tracker\n{text}")

    def _update_tray_tooltip(self, member: str, last: str) -> None:
        if self.tray is None:
            return
        line = f"Member: {member or 'not set'}   Last poll: {last or 'never'}"
        self.tray.setToolTip(f"ModDB Tracker\n{line}")

    def _on_worker_done(self, message: str) -> None:
        self._set_busy(False)
        self.refresh_all()
        self.set_status("Done.")
        if message:
            logger.info("Worker finished: %s", message)

    def _on_worker_failed(self, message: str) -> None:
        self._set_busy(False)
        self.refresh_all()
        self.set_status("Failed.")
        QMessageBox.critical(self, "Task failed", message)

    def _start_worker(self, fn: Callable, label: str) -> None:
        if self.worker is not None and self.worker.isRunning():
            return
        self._set_busy(True, label)
        self.worker = TrackerWorker(self.config, fn, parent=self)
        self.worker.done.connect(self._on_worker_done)
        self.worker.failed.connect(self._on_worker_failed)
        self.worker.start()

    def _poll_now(self) -> None:
        self._start_worker(run_poll, "Polling ModDB...")

    def _run_init(self) -> None:
        self._start_worker(run_init, "Setting up your account...")

    def _rescan(self) -> None:
        self._start_worker(run_discover, "Scanning profile...")

    def _regenerate_charts(self) -> None:
        def _charts(storage: Storage, config: Dict[str, Any]) -> Any:
            import charts

            paths = charts.generate_all(storage, Path(config["paths"]["output"]))
            return f"{len(paths)} chart(s) written"

        self._start_worker(_charts, "Regenerating charts...")

    # ---- reset ----------------------------------------------------------
    def _reset_app_data(self) -> None:
        reply = QMessageBox.question(
            self,
            "Delete all data",
            "This permanently deletes the database, all charts and logs, and resets config.json.\n\n"
            "The app will restart as a fresh first run. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            self._stopping = True
            self.auto_timer.stop()
            if self.worker is not None and self.worker.isRunning():
                self.worker.wait(5000)
            try:
                self.storage.close()
            except Exception:  # noqa: BLE001
                pass
            removed = reset_app_data(self.config)
            save_config(json.loads(json.dumps(DEFAULT_CONFIG)), self.config_path)
            logger.info("App data reset: %d file(s) removed", len(removed))
            QProcess.startDetached(sys.executable, [str(Path(__file__).resolve()), "--config", self.config_path])
            QMessageBox.information(
                self, "Data deleted",
                f"Removed {len(removed)} file(s). The app is restarting as a fresh first run.",
            )
            QApplication.instance().quit()
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Reset failed", str(exc))

    # ---- history backfill ----------------------------------------------
    def _backfill_mod(self, mod_id: int) -> None:
        mod = next((dict(m) for m in self.storage.get_mods(active_only=True) if int(m["id"]) == int(mod_id)), None)
        if mod is None:
            return

        def _run(storage: Storage, config: Dict[str, Any]) -> str:
            coverage = tracker.backfill_stats_history(storage, mod)
            return f"Backfilled {mod['name']}: {coverage['days']:,} day(s)"

        self._start_worker(_run, f"Backfilling {mod['name']}...")

    def _backfill_all(self) -> None:
        def _run(storage: Storage, config: Dict[str, Any]) -> str:
            return tracker.backfill_all_stats_history(storage, config)

        self._start_worker(_run, "Backfilling all mods...")

    # ---- per-mod actions ------------------------------------------------
    def _refresh_mod(self, mod_id: int) -> None:
        mod = next((dict(m) for m in self.storage.get_mods(active_only=True) if int(m["id"]) == int(mod_id)), None)
        if mod is None:
            return

        def _one(storage: Storage, config: Dict[str, Any]) -> str:
            target = next((dict(m) for m in storage.get_mods(active_only=True) if int(m["id"]) == int(mod_id)), None)
            if target is None:
                return "mod not found"
            tracker.snapshot_mod(storage, target, config, notify=False)
            return f"Refreshed {target['name']}"

        self._start_worker(_one, f"Refreshing {mod['name']}...")

    def _toggle_favorite(self, name_id: str, favorite: bool) -> None:
        self.storage.set_mod_favorite(name_id, favorite)
        self.refresh_all()

    def _remove_mod(self, name_id: str) -> None:
        mod = next((dict(m) for m in self.storage.get_mods(active_only=False) if str(m["name_id"]) == name_id), None)
        name = mod["name"] if mod else name_id
        answer = QMessageBox.question(
            self,
            "Remove from tracking",
            f"Stop tracking \u201c{name}\u201d?\nExisting history is kept, it just won't be polled anymore.",
        )
        if answer == QMessageBox.StandardButton.Yes:
            self.storage.set_mod_active(name_id, False)
            self.refresh_all()

    def _pick_save_path(self, title: str, default: str, flt: str) -> Optional[str]:
        path, _ = QFileDialog.getSaveFileName(self, title, default, flt)
        return path or None

    def _readonly_storage(self) -> Storage:
        return Storage(self.config["paths"]["db"])

    def _export_json(self) -> None:
        path = self._pick_save_path("Export data", "moddb_tracker_export.json", "JSON (*.json)")
        if not path:
            return
        try:
            storage = self._readonly_storage()
            try:
                storage.export_json(path)
            finally:
                storage.close()
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    def _export_csv(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "Choose a folder for CSV export")
        if not directory:
            return
        try:
            storage = self._readonly_storage()
            try:
                meta = {int(m["id"]): dict(m) for m in storage.get_mods(active_only=True)}
                csv_rows = []
                for m in storage.totals_per_mod():
                    mid = int(m["id"])
                    csv_rows.append([m["name"], m["downloads_total"], m["downloads_today"], m["visits"],
                                     m["watchers"], m["rating"], m["rank"],
                                     meta.get(mid, {}).get("content_type", "mod")])
                self._write_csv_rows(Path(directory) / "mods.csv",
                                     ["Name", "Downloads", "Today", "Visits", "Watchers", "Rating", "Rank", "Content type"],
                                     csv_rows)
                self._write_csv_rows(Path(directory) / "comments.csv",
                                     ["Posted", "Mod", "Author", "Text"],
                                     [list(r) for r in self._comment_list(storage)])
                self._write_csv_rows(Path(directory) / "events.csv",
                                     ["Time", "Kind", "Mod", "Message"],
                                     [[e["created_at"], e["kind"], e["mod_name"], e["message"]] for e in storage.recent_events(500)])
                self._write_csv_rows(Path(directory) / "history.csv",
                                     ["Mod", "Time", "Downloads", "Visits", "Watchers"],
                                     [[m["name"], s["fetched_at"], s["downloads_total"],
                                       s["visits"], s["watchers"]]
                                      for m in storage.get_mods(active_only=True)
                                      for s in storage.snapshots_for(int(m["id"]))][-2000:])
            finally:
                storage.close()
            QMessageBox.information(self, "CSV export", "Saved to:\n" + str(directory))
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    @staticmethod
    def _comment_list(storage) -> List[Tuple[str, str, str, Any]]:
        mod_names = {int(m["id"]): m["name"] for m in storage.get_mods(active_only=False)}
        return [(c["posted_at"], mod_names.get(int(c["mod_id"]), "-"), c["author"], c["content"])
                for c in storage.recent_comments(500)]

    @staticmethod
    def _write_csv_rows(path: Path, headers: List[str], rows: List[List[Any]]) -> None:
        import csv

        with open(str(path), "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            writer.writerows(rows)

    def _export_xlsx(self) -> None:
        path = self._pick_save_path("Export to Excel", "moddb_tracker.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            storage = self._readonly_storage()
            try:
                wb = Workbook()
                wb.remove(wb.active)
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="3B82F6")
                meta = {int(m["id"]): dict(m) for m in storage.get_mods(active_only=True)}
                xlsx_mods = []
                for m in storage.totals_per_mod():
                    mid = int(m["id"])
                    xlsx_mods.append([m["name"], m["downloads_total"], m["downloads_today"], m["visits"], m["watchers"],
                                      m["rating"], m["rank"], meta.get(mid, {}).get("content_type", "mod")])
                sections = {
                    "Mods": (["Name", "Downloads", "Today", "Visits", "Watchers", "Rating", "Rank", "Content type"],
                             xlsx_mods),
                    "Comments": (["Posted", "Mod", "Author", "Text"],
                                 [list(r) for r in self._comment_list(storage)]),
                    "Events": (["Time", "Kind", "Mod", "Message"],
                               [[e["created_at"], e["kind"], e["mod_name"], e["message"]] for e in storage.recent_events(500)]),
                    "History": (["Mod", "Time", "Downloads", "Visits", "Watchers"],
                                [[m["name"], s["fetched_at"], s["downloads_total"],
                                  s["visits"], s["watchers"]]
                                 for m in storage.get_mods(active_only=True)
                                 for s in storage.snapshots_for(int(m["id"]))][-2000:]),
                }
                for name, (headers, rows) in sections.items():
                    ws = wb.create_sheet(name)
                    ws.append(headers)
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                    for row in rows:
                        ws.append(row)
                    for col, _ in enumerate(headers, start=1):
                        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(12, min(45, len(str(headers[col - 1])) + 2))
                wb.save(path)
            finally:
                storage.close()
            QMessageBox.information(self, "Excel export", f"Saved to:\n{path}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    def _export_pdf(self) -> None:
        path = self._pick_save_path("Export PDF report", "moddb_tracker_report.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            storage = self._readonly_storage()
            try:
                summary = analytics.all_mods_summary(storage, 30)
                member = storage.meta_get("member_name") or "—"
                last_poll = storage.meta_get("last_poll") or "—"

                styles = getSampleStyleSheet()
                title_style = ParagraphStyle("TitleX", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#0F172A"))
                body = ParagraphStyle("BodyX", parent=styles["BodyText"], fontSize=9, leading=12)
                small = ParagraphStyle("SmallX", parent=styles["BodyText"], fontSize=8, textColor=colors.HexColor("#64748B"))

                doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
                                        topMargin=18 * mm, bottomMargin=18 * mm, title="ModDB Tracker report")
                story = [Paragraph("ModDB Tracker — weekly report", title_style),
                         Spacer(1, 4),
                         Paragraph(f"Member: {member} &nbsp;&nbsp;|&nbsp;&nbsp; Last poll: {last_poll} &nbsp;&nbsp;|&nbsp;&nbsp; "
                                   f"Generated: {datetime.datetime.now():%Y-%m-%d %H:%M}", small),
                         Spacer(1, 10)]

                agg = analytics.aggregate_summary(summary)
                story.append(Paragraph(
                    f"<b>Aggregate</b> — {agg['count']} tracked mods, {agg['total']:,} total downloads, "
                    f"<b>+{agg['delta_7d']:,}</b> in the last 7 days, <b>+{agg['delta_30d']:,}</b> in 30, "
                    f"~<b>{agg['next_week']:,}</b> projected for next week.", body))
                story.append(Spacer(1, 8))

                rows = [["Mod", "Total", "7d", "30d", "Avg/day", "Best day", "Next week"]]
                for s in summary:
                    best = s["best_day"]["label"] if s["best_day"] else "—"
                    nxt = f"{s['next_week_estimate']:,}" if s["next_week_estimate"] else "—"
                    rows.append([Paragraph(s["name"], body), f"{s['total']:,}", f"+{s['delta_7d']:,}",
                                 f"+{s['delta_30d']:,}", f"{s['avg_per_day']}", best, nxt])
                table = Table(rows, colWidths=[70 * mm, 22 * mm, 18 * mm, 18 * mm, 20 * mm, 28 * mm, 22 * mm], repeatRows=1)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]))
                story.append(table)
                story.append(Spacer(1, 10))

                comments = self._comment_list(storage)
                if comments:
                    story.append(Paragraph("<b>Recent comments</b>", body))
                    story.append(Spacer(1, 4))
                    for posted, mod_name, author, content in comments[:20]:
                        text = " ".join(str(content).split())
                        story.append(Paragraph(f"<b>{author}</b> on {mod_name} ({posted}): {text[:220]}", body))
                    story.append(Spacer(1, 8))

                events = storage.recent_events(20)
                if events:
                    story.append(Paragraph("<b>Recent events</b>", body))
                    story.append(Spacer(1, 4))
                    for e in events:
                        story.append(Paragraph(f"{e['created_at']} — {e['kind']} · {e['mod_name']}: {e['message']}", small))
                doc.build(story)
            finally:
                storage.close()
            QMessageBox.information(self, "PDF export", f"Saved to:\n{path}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    # ---- search ---------------------------------------------------------
    def _search_changed(self, text: str) -> None:
        needle = (text or "").strip()
        if len(needle) < 2:
            if self._search_override:
                self._search_override = False
                self.stack.setCurrentIndex(self._saved_page)
            return
        if not self._search_override:
            self._search_override = True
            self._saved_page = self.stack.currentIndex()
        self.search_page._storage = self.storage
        self.stack.setCurrentWidget(self.search_page)
        self.search_page.apply_filter(text)

    # ---- auto poll ------------------------------------------------------
    def _toggle_auto_poll(self, checked: bool) -> None:
        if checked:
            self.auto_timer.start(self._auto_interval_ms())
            self.set_status(f"Auto-poll every {self.config['poll']['interval_minutes']} min.")
        else:
            self.auto_timer.stop()
            self.set_status("Auto-poll off.")

    def _auto_interval_ms(self) -> int:
        return max(60, int(self.config["poll"]["interval_minutes"]) * 60 * 1000)

    def _restart_auto_timer(self) -> None:
        if self.auto_timer.isActive():
            self.auto_timer.start(self._auto_interval_ms())



def main() -> int:
    parser = argparse.ArgumentParser(description="ModDB Tracker GUI")
    parser.add_argument("--config", default=CONFIG_FILE, help="path to config.json")
    parser.add_argument("--minimized", action="store_true", help="start hidden in the system tray")
    args = parser.parse_args()

    app = QApplication(sys.argv)
    config = load_config(args.config)
    set_theme(str(config.get("ui", {}).get("theme", "dark")))
    app.setStyleSheet(build_qss(current_theme()))
    transport.patch_moddb()

    first_run = not is_profile_configured(config)
    account_set = False
    if first_run:
        dialog = FirstRunDialog()
        if dialog.exec() == FirstRunDialog.DialogCode.Accepted:
            config["profile_url"] = dialog.profile_url()
            config["auto_discover"] = True
            save_config(config, args.config)
            account_set = True

    window = TrackerWindow(args.config)

    if QSystemTrayIcon.isSystemTrayAvailable():
        app.setQuitOnLastWindowClosed(False)  # closing the window keeps the app in the tray
        if args.minimized or window.config.get("tray", {}).get("start_minimized", False):
            window.hide_to_tray(first_time=True)
        else:
            show_window(window)
    else:
        show_window(window)

    if first_run and account_set:
        QTimer.singleShot(700, window._run_init)
    elif window.config.get("ui", {}).get("poll_on_open", True) and is_profile_configured(window.config):
        QTimer.singleShot(700, window._poll_now)

    return app.exec()

