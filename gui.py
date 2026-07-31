"""PyQt6 GUI for the ModDB Tracker.

Shows charts, mod stats, download history, comments, events and lets you
run polls / rescans from the UI. Data tables live in the same SQLite DB the
CLI uses, so the GUI and the scheduled task are interchangeable.

Run with:  python gui.py [--config path/to/config.json]
"""

from __future__ import annotations

import argparse
import datetime
import json
import logging
import sys
import time
import traceback
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from PyQt6.QtCore import QObject, QRect, QSize, QThread, Qt, QTimer, QUrl, pyqtSignal
from PyQt6.QtGui import QAction, QColor, QDesktopServices, QFont, QIcon, QPainter, QPixmap
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLayout,
    QLayoutItem,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpinBox,
    QStackedWidget,
    QSystemTrayIcon,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

import transport
from storage import Storage

sys.path.insert(0, str(Path(__file__).resolve().parent))

import analytics  # noqa: E402
import tracker  # noqa: E402
from tracker import (  # noqa: E402  (imports depend on sys.path above)
    CONFIG_FILE,
    discover_mods,
    load_config,
    run_poll,
    save_config,
)

import pyqtgraph as pg  # noqa: E402
from pyqtgraph.exporters import ImageExporter  # noqa: E402

logger = logging.getLogger("tracker.gui")

VERSION = "2.1.0"

# --------------------------------------------------------------------------
# palette (dark blue theme)
# --------------------------------------------------------------------------

BG = "#0F172A"
CARD = "#1E293B"
PANEL = "#1E293B"
PANEL2 = "#243247"
SURFACE = "#273449"
BORDER = "#334155"
TEXT = "#E2E8F0"
GRAY = "#94A3B8"
FAINT = "#64748B"
ACCENT = "#3B82F6"
ACCENT_DARK = "#2563EB"
SUCCESS = "#22C55E"
WARNING = "#F59E0B"
ERROR = "#EF4444"


# --------------------------------------------------------------------------
# tray icon
# --------------------------------------------------------------------------

def tray_icon_path() -> str:
    """Return a cached .png for the tray icon, drawing it if missing."""
    path = Path(__file__).resolve().parent / "tray_icon.png"
    if path.exists():
        return str(path)
    pixmap = QPixmap(64, 64)
    pixmap.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    painter.setBrush(QColor(ACCENT))
    painter.setPen(Qt.PenStyle.NoPen)
    painter.drawRoundedRect(4, 4, 56, 56, 14, 14)
    painter.setPen(QColor("#FFFFFF"))
    font = QFont("Segoe UI", 22, QFont.Weight.Bold)
    painter.setFont(font)
    painter.drawText(pixmap.rect(), Qt.AlignmentFlag.AlignCenter, "M")
    painter.end()
    pixmap.save(str(path))
    return str(path)


# --------------------------------------------------------------------------
# global stylesheet
# --------------------------------------------------------------------------

QSS = f"""
* {{
    outline: none;
}}
QWidget {{
    background-color: {BG};
    color: {TEXT};
    font-family: "Segoe UI";
    font-size: 13px;
}}
QLabel {{ background: transparent; }}

QFrame#Panel, QFrame#PanelSecondary {{
    background-color: {CARD};
    border: 1px solid {BORDER};
    border-radius: 10px;
}}
QFrame#TopBar {{
    background-color: {BG};
    border-bottom: 1px solid {BORDER};
    border-radius: 0px;
}}
QFrame#StatCard, QFrame#ModCard {{
    background-color: {CARD};
    border: 1px solid {BORDER};
    border-radius: 12px;
}}
QFrame#ModCard:hover {{
    border: 1px solid {ACCENT};
    background-color: {SURFACE};
}}
QFrame#EventRow {{
    background-color: transparent;
    border: none;
}}
QFrame#EventRow:hover {{ background-color: {PANEL2}; border-radius: 6px; }}

QPushButton {{
    background-color: {PANEL2};
    border: 1px solid {BORDER};
    border-radius: 7px;
    padding: 7px 14px;
    font-weight: 600;
}}
QPushButton:hover {{ border-color: {ACCENT}; background-color: {SURFACE}; }}
QPushButton:pressed {{ background-color: {BORDER}; }}
QPushButton:disabled {{ color: {FAINT}; border-color: {BORDER}; background-color: {PANEL2}; }}
QPushButton#Primary {{
    background-color: {ACCENT};
    border: none;
    color: #FFFFFF;
}}
QPushButton#Primary:hover {{ background-color: {ACCENT_DARK}; }}
QPushButton#Primary:disabled {{ background-color: {PANEL2}; color: {FAINT}; }}
QPushButton#Danger:hover {{ border-color: {ERROR}; color: {ERROR}; }}

QToolButton {{
    background: transparent;
    border: none;
    border-radius: 6px;
    padding: 4px 8px;
    font-weight: 600;
}}
QToolButton:hover {{ background-color: {PANEL2}; color: {ACCENT}; }}

QLineEdit, QSpinBox, QComboBox, QPlainTextEdit, QTextEdit {{
    background-color: {PANEL2};
    border: 1px solid {BORDER};
    border-radius: 8px;
    padding: 7px 10px;
    selection-background-color: {ACCENT};
}}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus, QPlainTextEdit:focus {{
    border: 1px solid {ACCENT};
}}
QLineEdit#Search {{
    border-radius: 18px;
    padding: 7px 16px;
    background-color: {PANEL2};
    border: 1px solid {BORDER};
}}
QLineEdit#Search:focus {{ border: 1px solid {ACCENT}; }}

QComboBox::drop-down {{ border: none; width: 24px; }}
QComboBox::down-arrow {{
    image: none;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 6px solid {GRAY};
    margin-right: 10px;
}}
QComboBox QAbstractItemView {{
    background-color: {PANEL2};
    border: 1px solid {BORDER};
    border-radius: 8px;
    selection-background-color: {ACCENT};
    selection-color: #FFFFFF;
}}

QCheckBox {{ spacing: 8px; }}
QCheckBox::indicator {{
    width: 17px;
    height: 17px;
    border: 2px solid {BORDER};
    border-radius: 4px;
    background-color: {PANEL2};
}}
QCheckBox::indicator:hover {{ border-color: {ACCENT}; }}
QCheckBox::indicator:checked {{
    background-color: {ACCENT};
    border-color: {ACCENT};
    image: none;
}}

QTableWidget {{
    background-color: {CARD};
    alternate-background-color: {PANEL2};
    border: 1px solid {BORDER};
    border-radius: 10px;
    gridline-color: transparent;
}}
QTableWidget::item {{ padding: 6px 8px; border: none; }}
QTableWidget::item:selected {{ background-color: {ACCENT}; color: #FFFFFF; }}
QHeaderView::section {{
    background-color: {SURFACE};
    color: {GRAY};
    border: none;
    border-bottom: 1px solid {BORDER};
    padding: 9px 8px;
    font-weight: 700;
    font-size: 11px;
    letter-spacing: 1px;
}}

QListWidget#Sidebar {{
    background-color: {BG};
    border: none;
    border-radius: 10px;
}}
QListWidget#Sidebar::item {{
    border-radius: 8px;
    padding: 9px 12px;
    margin: 2px 6px;
    color: {GRAY};
}}
QListWidget#Sidebar::item:hover {{
    background-color: {PANEL2};
    color: {TEXT};
}}
QListWidget#Sidebar::item:selected {{
    background-color: {ACCENT};
    color: #FFFFFF;
    font-weight: 700;
}}
QListWidget {{
    background-color: {CARD};
    border: 1px solid {BORDER};
    border-radius: 8px;
}}
QListWidget::item {{ padding: 4px 6px; }}

QScrollBar:vertical {{
    background: transparent;
    width: 10px;
    margin: 0;
}}
QScrollBar::handle:vertical {{
    background: {BORDER};
    border-radius: 5px;
    min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{ background: {FAINT}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: transparent;
    height: 10px;
    margin: 0;
}}
QScrollBar::handle:horizontal {{
    background: {BORDER};
    border-radius: 5px;
    min-width: 30px;
}}
QScrollBar::handle:horizontal:hover {{ background: {FAINT}; }}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}

QProgressBar {{
    background-color: {PANEL2};
    border: 1px solid {BORDER};
    border-radius: 5px;
    height: 10px;
    text-align: center;
    color: transparent;
}}
QProgressBar::chunk {{
    background-color: {ACCENT};
    border-radius: 4px;
}}

QLabel#PageTitle {{
    font-size: 22px;
    font-weight: 800;
    letter-spacing: 0.5px;
    background: transparent;
}}
QLabel#SectionTitle {{
    font-size: 11px;
    font-weight: 800;
    letter-spacing: 2px;
    color: {GRAY};
    background: transparent;
}}
QLabel#Hint {{
    font-size: 12px;
    color: {GRAY};
    background: transparent;
}}
QLabel#StatCaption {{
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.5px;
    color: {GRAY};
}}
QLabel#StatValue {{
    font-size: 26px;
    font-weight: 800;
    color: {TEXT};
    background: transparent;
}}
QLabel#StatDelta {{
    font-size: 12px;
    color: {GRAY};
    background: transparent;
}}
QMenu {{
    background-color: {PANEL2};
    border: 1px solid {BORDER};
    border-radius: 8px;
    padding: 6px;
}}
QMenu::item {{ padding: 6px 22px; border-radius: 5px; }}
QMenu::item:selected {{ background-color: {ACCENT}; color: #FFFFFF; }}
QMenu::separator {{
    height: 1px;
    background: {BORDER};
    margin: 5px 8px;
}}
QToolTip {{
    background-color: {SURFACE};
    color: {TEXT};
    border: 1px solid {BORDER};
    padding: 6px;
    border-radius: 4px;
}}
QStatusBar {{
    background: {BG};
    color: {GRAY};
    border-top: 1px solid {BORDER};
}}
QStatusBar::item {{ border: none; }}
"""


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------

def relative_time(iso_str: Any) -> str:
    """Human-friendly time like '2h ago', 'just now'."""
    try:
        now = datetime.datetime.now().astimezone()
        ts = datetime.datetime.fromisoformat(str(iso_str))
        if ts.tzinfo is None:
            ts = ts.astimezone()
        diff = now - ts
        secs = int(diff.total_seconds())
        if secs < 60:
            return "just now"
        if secs < 3600:
            return f"{secs // 60}m ago"
        if secs < 86400:
            return f"{secs // 3600}h ago"
        if secs < 86400 * 7:
            return f"{secs // 86400}d ago"
        return ts.strftime("%b %d, %Y")
    except Exception:  # noqa: BLE001
        return str(iso_str)[:16]


def format_num(value: Any) -> str:
    try:
        return f"{int(value or 0):,}"
    except Exception:  # noqa: BLE001
        return str(value or "—")


class FlowLayout(QLayout):
    """Responsive grid: cards wrap into columns that fit the available width."""

    def __init__(self, parent: Optional[QWidget] = None, spacing: int = 12, min_width: int = 300) -> None:
        super().__init__(parent)
        self.setSpacing(spacing)
        self.setContentsMargins(0, 0, 0, 0)
        self._items: List[QLayoutItem] = []
        self._min_width = min_width
        self._spacing = spacing

    def addItem(self, item) -> None:  # noqa: D102
        self._items.append(item)

    def count(self) -> int:  # noqa: D102
        return len(self._items)

    def itemAt(self, index: int):  # noqa: D102
        if 0 <= index < len(self._items):
            return self._items[index]
        return None

    def takeAt(self, index: int):  # noqa: D102
        if 0 <= index < len(self._items):
            return self._items.pop(index)
        return None

    def expandingDirections(self):  # noqa: D102
        return Qt.Orientation(0)

    def hasHeightForWidth(self) -> bool:  # noqa: D102
        return True

    def heightForWidth(self, width: int) -> int:  # noqa: D102
        return self._measure(width)

    def minimumHeightForWidth(self, width: int) -> int:  # noqa: D102
        return self._measure(width)

    def sizeHint(self) -> QSize:  # noqa: D102
        return QSize(self._min_width, self._measure(self._min_width))

    def minimumSize(self) -> QSize:  # noqa: D102
        return QSize(self._min_width, 0)

    def add_card(self, widget: QWidget) -> None:
        self.addWidget(widget)
        widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

    def clear(self) -> None:
        while self._items:
            item = self.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

    def _columns_for(self, width: int) -> int:
        if width < self._min_width:
            return 1
        return max(1, width // self._min_width)

    def _measure(self, width: int) -> int:
        if not self._items:
            return 0
        heights = [it.sizeHint().height() for it in self._items if it.widget()]
        if not heights:
            return 0
        h = max(heights)
        cols = self._columns_for(width)
        rows = (len(heights) + cols - 1) // cols
        return rows * h + (rows - 1) * self._spacing

    def setGeometry(self, rect) -> None:  # noqa: N802
        super().setGeometry(rect)
        if not self._items:
            return
        items = [it for it in self._items if it.widget()]
        width = rect.width()
        cols = self._columns_for(width)
        item_w = (width - (cols - 1) * self._spacing) // cols
        for i, it in enumerate(items):
            col = i % cols
            row = i // cols
            x = rect.x() + col * (item_w + self._spacing)
            y = rect.y() + row * (it.sizeHint().height() + self._spacing)
            it.setGeometry(QRect(x, y, item_w, it.sizeHint().height()))


# --------------------------------------------------------------------------
# background worker
# --------------------------------------------------------------------------

class TrackerWorker(QThread):
    done = pyqtSignal(str)
    failed = pyqtSignal(str)

    def __init__(self, config: Dict[str, Any], fn: Callable, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._config = config
        self._fn = fn

    def run(self) -> None:  # noqa: D102
        try:
            storage = Storage(self._config["paths"]["db"])
            try:
                message = self._fn(storage, self._config) or ""
                if not isinstance(message, str):
                    message = str(message)
                self.done.emit(message)
            finally:
                storage.close()
        except Exception as exc:  # noqa: BLE001
            traceback.print_exc()
            self.failed.emit(str(exc))


def run_discover(storage: Storage, config: Dict[str, Any]) -> Any:
    """Rescan the member profile for mods/addons/files (runs on the worker)."""
    from tracker import discover_mods

    found = discover_mods(storage, config)
    return f"Found {len(found)} item(s)"


# --------------------------------------------------------------------------
# logging bridge
# --------------------------------------------------------------------------

class LogBridge(QObject):
    line = pyqtSignal(str)


class GuiLogHandler(logging.Handler):
    def __init__(self, bridge: LogBridge) -> None:
        super().__init__()
        self._bridge = bridge

    def emit(self, record: logging.LogRecord) -> None:  # noqa: D102
        try:
            message = self.format(record)
            self._bridge.line.emit(message)
        except Exception:  # noqa: BLE001
            pass


# --------------------------------------------------------------------------
# reusable widgets
# --------------------------------------------------------------------------

def panel(parent: QWidget) -> QFrame:
    frame = QFrame(parent)
    frame.setObjectName("Panel")
    return frame


def section_label(text: str) -> QLabel:
    label = QLabel(text.upper())
    label.setObjectName("SectionTitle")
    return label


def make_table(headers: List[str], selectable: bool = True) -> QTableWidget:
    table = QTableWidget(0, len(headers))
    table.setHorizontalHeaderLabels(headers)
    table.verticalHeader().setVisible(False)
    table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
    table.setAlternatingRowColors(True)
    table.setShowGrid(False)
    table.setWordWrap(False)
    if selectable:
        table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
    else:
        table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
    header = table.horizontalHeader()
    header.setStretchLastSection(True)
    header.setHighlightSections(False)
    table.setSortingEnabled(False)
    return table


def fill_table(table: QTableWidget, rows: List[List[Any]]) -> None:
    table.setSortingEnabled(False)
    table.setRowCount(0)
    table.setRowCount(len(rows))
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            item = QTableWidgetItem("" if value is None else str(value))
            if c == 0:
                item.setData(Qt.ItemDataRole.UserRole, value)
            table.setItem(r, c, item)
    table.resizeColumnsToContents()


class StatCard(QFrame):
    """Dashboard stat card with count-up animation and colored delta."""

    def __init__(self, title: str, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("StatCard")
        self.setMinimumHeight(98)

        self._title = QLabel(str(title).upper())
        self._title.setObjectName("StatCaption")
        self._value = QLabel("—")
        self._value.setObjectName("StatValue")
        self._delta = QLabel("")
        self._delta.setObjectName("StatDelta")

        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 13, 16, 13)
        lay.setSpacing(2)
        lay.addWidget(self._title)
        lay.addWidget(self._value)
        lay.addWidget(self._delta)

        self._current = 0
        self._target = 0
        self._start = 0
        self._t0 = 0.0
        self._dur = 0.6
        self._timer = QTimer(self)
        self._timer.setInterval(30)
        self._timer.timeout.connect(self._step)

    def set_value(self, target: Any, delta_text: str = "", delta_color: Optional[str] = None) -> None:
        try:
            target = int(float(target))
        except Exception:  # noqa: BLE001
            target = 0
        self._delta.setText(delta_text)
        self._delta.setStyleSheet(f"color: {delta_color or GRAY}; font-size: 12px;")
        self._target = target
        self._start = self._current
        self._t0 = time.monotonic()
        self._timer.start()

    def set_text(self, text: str, delta_text: str = "", delta_color: Optional[str] = None) -> None:
        self._timer.stop()
        self._value.setText(str(text))
        self._delta.setText(delta_text)
        self._delta.setStyleSheet(f"color: {delta_color or GRAY}; font-size: 12px;")

    def _step(self) -> None:
        frac = min(1.0, (time.monotonic() - self._t0) / self._dur)
        eased = 1 - (1 - frac) ** 3
        val = int(round(self._start + (self._target - self._start) * eased))
        self._value.setText(f"{val:,}")
        if frac >= 1.0:
            self._timer.stop()
            self._current = self._target


class ChartCard(QFrame):
    """A panel showing one chart PNG, rescaled to fit its column."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("Panel")
        self._pixmap: Optional[QPixmap] = None
        self._label = QLabel()
        self._label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._label.setMinimumHeight(180)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.addWidget(self._label)

    def set_image(self, path: Path) -> None:
        self._pixmap = QPixmap(str(path))
        self._path = path
        self.rescale(720)

    def rescale(self, width: int) -> None:
        if self._pixmap is None or self._pixmap.isNull():
            self._label.setText(str(self._path.name))
            return
        h = max(140, int(self._pixmap.height() * width / max(1, self._pixmap.width())))
        self._label.setPixmap(self._pixmap.scaled(
            max(120, width), h,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        ))


def _date_to_ts(d: datetime.date) -> float:
    return datetime.datetime.combine(d, datetime.time()).timestamp()


class PlotCard(QFrame):
    """Panel with an interactive pyqtgraph chart and a hover readout line."""

    def __init__(self, title: str, subtitle: str, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("Panel")
        self._series: List[Dict[str, Any]] = []

        head = QLabel(title)
        head.setObjectName("SectionTitle")
        sub = QLabel(subtitle)
        sub.setObjectName("Caption")

        self.plot = pg.PlotWidget(axisItems={"bottom": pg.DateAxisItem()})
        self.plot.setBackground(CARD)
        self.plot.setMinimumHeight(230)
        pi = self.plot.getPlotItem()
        pi.showGrid(x=True, y=True, alpha=0.18)
        pi.setMouseEnabled(x=True, y=False)
        for axis_name in ("left", "bottom"):
            axis = pi.getAxis(axis_name)
            axis.setPen(pg.mkPen(BORDER))
            axis.setTextPen(pg.mkPen(FAINT))
            try:
                axis.setTickFont(QFont("Segoe UI", 9))
            except Exception:  # noqa: BLE001
                pass

        self.info = QLabel("Hover for values  \u00b7  drag to zoom")
        self.info.setObjectName("Caption")

        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 10)
        lay.setSpacing(8)
        lay.addWidget(head)
        lay.addWidget(sub)
        lay.addWidget(self.plot, 1)
        lay.addWidget(self.info)

        self.plot.scene().sigMouseMoved.connect(self._on_hover)

    # ---- drawing helpers ------------------------------------------------
    def clear_chart(self) -> None:
        self.plot.clear()
        self._series = []

    def set_ylabel(self, text: str) -> None:
        self.plot.getPlotItem().setLabel("left", text)

    def add_bars(self, dates, values, color: str, name: str = "", alpha: int = 255) -> None:
        ts = [_date_to_ts(d) for d in dates]
        brush = QColor(color)
        brush.setAlpha(alpha)
        item = pg.BarGraphItem(
            x=ts, height=values, width=86400 * 0.68,
            brush=pg.mkBrush(brush), pen=pg.mkPen(color),
        )
        self.plot.addItem(item)
        self._series.append({"times": ts, "values": list(values), "labels": list(dates), "name": name})

    def add_line(self, dates, values, color: str, width: int = 2, fill: bool = False, name: str = "") -> None:
        ts = [_date_to_ts(d) for d in dates]
        item = self.plot.plot(
            ts, values, pen=pg.mkPen(color, width=width), antialias=True,
        )
        if fill:
            item.setFillLevel(0)
            brush = QColor(color)
            brush.setAlpha(45)
            item.setBrush(pg.mkBrush(brush))
        self._series.append({"times": ts, "values": list(values), "labels": list(dates), "name": name})

    def add_milestone(self, date: datetime.date, threshold: int) -> None:
        ts = _date_to_ts(date)
        self.plot.addItem(pg.InfiniteLine(
            pos=ts, angle=90, pen=pg.mkPen(WARNING, width=1, style=Qt.PenStyle.DashLine),
        ))
        text = pg.TextItem(f"{threshold:,}", color=WARNING, anchor=(1, 1))
        text.setPos(ts, threshold)
        self.plot.addItem(text)

    def _on_hover(self, pos) -> None:
        vb = self.plot.getPlotItem().getViewBox()
        if not vb.sceneBoundingRect().contains(pos):
            return
        mouse = vb.mapSceneToView(pos)
        x = mouse.x()
        best = None
        best_dist = None
        for entry in self._series:
            times = entry["times"]
            if not times:
                continue
            idx = min(range(len(times)), key=lambda i: abs(times[i] - x))
            dist = abs(times[idx] - x)
            if best_dist is None or dist < best_dist:
                best_dist = dist
                best = (entry, idx)
        if best is None:
            self.info.setText("")
            return
        entry, idx = best
        label = entry["labels"][idx]
        if isinstance(label, datetime.date):
            label = label.strftime("%Y-%m-%d")
        prefix = f"{entry['name']}   " if entry.get("name") else ""
        self.info.setText(f"{prefix}{label}   \u2192   {entry['values'][idx]:,}")


class AnalyticsPage(QWidget):
    """Interactive downloads analytics: pyqtgraph charts with zoom, hover, export."""

    LINE_COLORS = [ACCENT, SUCCESS, WARNING, "#A78BFA", "#38BDF8", "#F472B6", "#FB923C", "#34D399"]

    def __init__(self, config: Dict[str, Any], config_path: str = CONFIG_FILE, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._config = config
        self._config_path = config_path
        self._storage: Optional[Storage] = None
        self._summaries: List[Dict[str, Any]] = []

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(0, 0, 8, 0)
        layout.setSpacing(12)

        title = QLabel("Analytics")
        title.setObjectName("PageTitle")
        layout.addWidget(title)
        hint = QLabel("Interactive downloads analytics \u2014 hover a chart for values, drag to zoom.")
        hint.setObjectName("Hint")
        layout.addWidget(hint)

        controls = QHBoxLayout()
        controls.setSpacing(10)
        controls.addWidget(QLabel("Mod:"))
        self.mod_combo = QComboBox()
        self.mod_combo.setMinimumWidth(240)
        self.mod_combo.currentIndexChanged.connect(self._mod_changed)
        controls.addWidget(self.mod_combo)
        controls.addWidget(QLabel("Range:"))
        self.days_combo = QComboBox()
        for n in (30, 60, 90):
            self.days_combo.addItem(f"Last {n} days", n)
        self.days_combo.setCurrentIndex(self._default_days_index())
        self.days_combo.currentIndexChanged.connect(self._days_changed)
        controls.addWidget(self.days_combo)
        controls.addStretch(1)
        self.export_btn = QPushButton("Export charts\u2026")
        self.export_btn.clicked.connect(self._export_charts)
        controls.addWidget(self.export_btn)
        layout.addLayout(controls)

        cards = QHBoxLayout()
        cards.setSpacing(12)
        self.stat_total = StatCard("Total downloads")
        self.stat_7d = StatCard("Last 7 days")
        self.stat_30d = StatCard("Last 30 days")
        self.stat_next = StatCard("Next 7 days (est.)")
        for c in (self.stat_total, self.stat_7d, self.stat_30d, self.stat_next):
            cards.addWidget(c, 1)
        layout.addLayout(cards)

        self.highlights = QFrame()
        self.highlights.setObjectName("Panel")
        self.highlights_label = QLabel("")
        self.highlights_label.setWordWrap(True)
        hl = QHBoxLayout(self.highlights)
        hl.setContentsMargins(14, 10, 14, 10)
        hl.addWidget(self.highlights_label)
        layout.addWidget(self.highlights)

        self.grid_host = QWidget()
        grid = QGridLayout(self.grid_host)
        grid.setSpacing(12)
        self.plot_daily = PlotCard("Downloads per day", "Bars: daily gain   \u00b7   line: 7-day average")
        self.plot_cum = PlotCard("Cumulative downloads", "Total downloads over time")
        self.plot_weekly = PlotCard("Weekly downloads", "Downloads gained per ISO week")
        grid.addWidget(self.plot_daily, 0, 0)
        grid.addWidget(self.plot_cum, 0, 1)
        grid.addWidget(self.plot_weekly, 1, 0, 1, 2)
        layout.addWidget(self.grid_host, 1)

        self.placeholder = QLabel("No data yet. Run a poll to start collecting download history.")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setObjectName("Hint")
        self.placeholder.setMinimumHeight(160)
        layout.addWidget(self.placeholder, 1)

        self.scroll.setWidget(content)
        outer.addWidget(self.scroll)

    def _default_days_index(self) -> int:
        days = int(self._config.get("ui", {}).get("analytics_days", 60))
        idx = self.days_combo.findData(days)
        return idx if idx >= 0 else 1

    # ---- refresh --------------------------------------------------------
    def refresh(self, storage: Storage) -> None:
        self._storage = storage
        days = int(self.days_combo.currentData() or 60)
        self._summaries = analytics.all_mods_summary(storage, days)
        current = self.mod_combo.currentData()
        self.mod_combo.blockSignals(True)
        self.mod_combo.clear()
        self.mod_combo.addItem("All tracked mods", None)
        for s in self._summaries:
            self.mod_combo.addItem(s["name"], s["mod_id"])
        if current is not None:
            idx = self.mod_combo.findData(current)
            if idx >= 0:
                self.mod_combo.setCurrentIndex(idx)
        self.mod_combo.blockSignals(False)
        self._update()

    def _days_changed(self) -> None:
        try:
            self._config.setdefault("ui", {})["analytics_days"] = int(self.days_combo.currentData() or 60)
            save_config(self._config, self._config_path)
        except Exception:  # noqa: BLE001
            pass
        if self._storage is not None:
            self.refresh(self._storage)

    def _mod_changed(self) -> None:
        self._update()

    # ---- rendering ------------------------------------------------------
    def _update(self) -> None:
        has = bool(self._summaries)
        self.placeholder.setVisible(not has)
        self.grid_host.setVisible(has)
        self.highlights.setVisible(has)
        for card in (self.plot_daily, self.plot_cum, self.plot_weekly):
            card.clear_chart()
        if not has:
            for s in (self.stat_total, self.stat_7d, self.stat_30d, self.stat_next):
                s.set_text("—")
            return

        sel = self.mod_combo.currentData()
        if sel is None:
            self._update_all()
        else:
            self._update_mod(sel)

    def _fill_cards(self, total: int, d7: int, d30: int, nxt: Optional[int]) -> None:
        self.stat_total.set_value(total, f"last 30 days: +{d30:,}", SUCCESS if d30 else GRAY)
        self.stat_7d.set_text(f"{d7:,}", "last 7 days", SUCCESS)
        self.stat_30d.set_text(f"{d30:,}", "last 30 days", SUCCESS)
        if nxt:
            self.stat_next.set_text(f"{nxt:,}", "linear projection", ACCENT)
        else:
            self.stat_next.set_text("—", "not enough data", GRAY)

    def _highlights_line(self, parts: List[str]) -> None:
        self.highlights_label.setText("  \u00b7  ".join(parts))
        self.highlights_label.setStyleSheet(f"color: {GRAY}; font-size: 12px;")

    def _update_all(self) -> None:
        agg = analytics.aggregate_summary(self._summaries)
        self._fill_cards(agg["total"], agg["delta_7d"], agg["delta_30d"], agg["next_week"])

        top = self._summaries[0]
        fastest = max(self._summaries, key=lambda s: s["delta_7d"]) if self._summaries else None
        ranked = ", ".join(f"{i + 1}. {s['name']} ({s['total']:,})" for i, s in enumerate(self._summaries[:3]))
        parts = [f"{agg['count']} tracked mods",
                 f"top: {top['name']} ({top['total']:,})",
                 f"fastest 7d: {fastest['name']} (+{fastest['delta_7d']:,})" if fastest else None,
                 f"ranking: {ranked}"]
        self._highlights_line([p for p in parts if p])

        today = datetime.date.today()
        days = int(self.days_combo.currentData() or 60)
        start = today - datetime.timedelta(days=days - 1)

        # daily deltas summed across mods
        agg_daily: Dict[datetime.date, int] = {}
        for s in self._summaries:
            for d, v in s["deltas"]:
                agg_daily[d] = agg_daily.get(d, 0) + v
        dailies = sorted(agg_daily.items())
        self.plot_daily.set_ylabel("Downloads")
        if dailies:
            self.plot_daily.add_bars(*zip(*dailies), ACCENT)
            ma = analytics.moving_average(dailies, 7)
            self.plot_daily.add_line(*zip(*ma), "#38BDF8", width=2)

        # cumulative per mod, aligned to the same window
        for i, s in enumerate(self._summaries):
            series = [p for p in s["series"] if p[0] >= start]
            totals = analytics.daily_totals_range(series, days)
            color = self.LINE_COLORS[i % len(self.LINE_COLORS)]
            self.plot_cum.add_line(*zip(*totals), color, width=2)
        self.plot_cum.set_ylabel("Total downloads")

        # weekly totals across mods
        weeks: Dict[datetime.date, int] = {}
        for s in self._summaries:
            for d, v in s["weeks"]:
                weeks[d] = weeks.get(d, 0) + v
        wk = sorted(weeks.items())
        self.plot_weekly.set_ylabel("Downloads")
        if wk:
            self.plot_weekly.add_bars(*zip(*wk), SUCCESS)

    def _update_mod(self, mod_id: int) -> None:
        s = next((x for x in self._summaries if x["mod_id"] == mod_id), None)
        if s is None:
            return
        self._fill_cards(s["total"], s["delta_7d"], s["delta_30d"], s["next_week_estimate"])

        parts = [f"first seen: {s['first_seen']}",
                 f"best day: {s['best_day']['label']} (+{s['best_day']['value']:,})" if s["best_day"] else None,
                 f"best week: {s['best_week']['label']} (+{s['best_week']['value']:,})" if s["best_week"] else None,
                 f"avg/day: {s['avg_per_day']}"]
        if s["milestones"]:
            ms = ", ".join(f"{m['threshold']:,} ({m['date']})" for m in s["milestones"])
            parts.append(f"milestones: {ms}")
        else:
            parts.append("milestones: none reached in range")
        self._highlights_line([p for p in parts if p])

        self.plot_daily.set_ylabel("Downloads")
        if s["deltas"]:
            self.plot_daily.add_bars(*zip(*s["deltas"]), ACCENT)
            self.plot_daily.add_line(*zip(*s["ma7"]), "#38BDF8", width=2)

        totals = analytics.daily_totals_range(s["series"], s["days"])
        self.plot_cum.set_ylabel("Total downloads")
        self.plot_cum.add_line(*zip(*totals), ACCENT, width=2, fill=True)
        for m in s["milestones"]:
            self.plot_cum.add_milestone(m["date"], m["threshold"])

        self.plot_weekly.set_ylabel("Downloads")
        if s["weeks"]:
            self.plot_weekly.add_bars(*zip(*s["weeks"]), SUCCESS)

    # ---- export ---------------------------------------------------------
    def _export_charts(self) -> None:
        dir_path = QFileDialog.getExistingDirectory(self, "Choose a folder for chart PNGs")
        if not dir_path:
            return
        try:
            written = []
            for slug, card in (("daily", self.plot_daily), ("cumulative", self.plot_cum), ("weekly", self.plot_weekly)):
                path = Path(dir_path) / f"analytics_{slug}.png"
                exporter = ImageExporter(card.plot.getPlotItem())
                exporter.parameters()["width"] = 1200
                exporter.export(str(path))
                written.append(path.name)
            QMessageBox.information(self, "Charts exported", "Saved:\n" + "\n".join(written))
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))


class ComparePage(QWidget):
    """Head-to-head comparison of two tracked mods (Phase 3)."""

    COLOR_A = ACCENT
    COLOR_B = SUCCESS

    def __init__(self, config: Dict[str, Any], config_path: str = CONFIG_FILE, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._config = config
        self._config_path = config_path
        self._storage: Optional[Storage] = None
        self._mods: List[Dict[str, Any]] = []
        self._summary_a: Optional[Dict[str, Any]] = None
        self._summary_b: Optional[Dict[str, Any]] = None

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(0, 0, 8, 0)
        layout.setSpacing(12)

        title = QLabel("Compare")
        title.setObjectName("PageTitle")
        layout.addWidget(title)
        hint = QLabel("Compare two mods head-to-head \u2014 totals, growth and the daily advantage.")
        hint.setObjectName("Hint")
        layout.addWidget(hint)

        controls = QHBoxLayout()
        controls.setSpacing(10)
        self.combo_a = QComboBox()
        self.combo_a.setMinimumWidth(230)
        self.combo_a.currentIndexChanged.connect(self._on_combo_changed)
        self.combo_b = QComboBox()
        self.combo_b.setMinimumWidth(230)
        self.combo_b.currentIndexChanged.connect(self._on_combo_changed)
        controls.addWidget(QLabel("Mod A:"))
        controls.addWidget(self.combo_a)
        controls.addWidget(QLabel("Mod B:"))
        controls.addWidget(self.combo_b)
        controls.addWidget(QLabel("Range:"))
        self.days_combo = QComboBox()
        for n in (30, 60, 90):
            self.days_combo.addItem(f"Last {n} days", n)
        self.days_combo.setCurrentIndex(1)
        self.days_combo.currentIndexChanged.connect(self._days_changed)
        controls.addWidget(self.days_combo)
        controls.addStretch(1)
        self.export_btn = QPushButton("Export charts\u2026")
        self.export_btn.clicked.connect(self._export_charts)
        controls.addWidget(self.export_btn)
        layout.addLayout(controls)

        self.table = make_table(["Metric", "Mod A", "Mod B"], selectable=False)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.setMinimumHeight(330)
        self.table.setMaximumHeight(360)
        layout.addWidget(self.table)

        self.grid_host = QWidget()
        grid = QGridLayout(self.grid_host)
        grid.setSpacing(12)
        self.plot_cum = PlotCard("Cumulative downloads", f"blue = Mod A   \u00b7   green = Mod B")
        self.plot_daily = PlotCard("Downloads per day", "Daily gain, both mods")
        self.plot_diff = PlotCard("Daily advantage", "green: A ahead   \u00b7   red: B ahead")
        grid.addWidget(self.plot_cum, 0, 0)
        grid.addWidget(self.plot_daily, 0, 1)
        grid.addWidget(self.plot_diff, 1, 0, 1, 2)
        layout.addWidget(self.grid_host, 1)

        self.placeholder = QLabel("Track at least two mods to compare them here.")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setObjectName("Hint")
        self.placeholder.setMinimumHeight(160)
        layout.addWidget(self.placeholder, 1)

        self.scroll.setWidget(content)
        outer.addWidget(self.scroll)

    # ---- refresh --------------------------------------------------------
    def refresh(self, storage: Storage) -> None:
        self._storage = storage
        self._mods = [dict(m) for m in storage.get_mods(active_only=True)]
        current_a = self.combo_a.currentData()
        current_b = self.combo_b.currentData()
        self.combo_a.blockSignals(True)
        self.combo_b.blockSignals(True)
        self.combo_a.clear()
        self.combo_b.clear()
        for m in self._mods:
            self.combo_a.addItem(m["name"], int(m["id"]))
            self.combo_b.addItem(m["name"], int(m["id"]))
        if len(self._mods) >= 2:
            if current_a is not None and self.combo_a.findData(current_a) >= 0:
                self.combo_a.setCurrentIndex(self.combo_a.findData(current_a))
            elif len(self._mods) >= 2:
                self.combo_a.setCurrentIndex(0)
            if current_b is not None and self.combo_b.findData(current_b) >= 0:
                self.combo_b.setCurrentIndex(self.combo_b.findData(current_b))
            elif len(self._mods) >= 2:
                self.combo_b.setCurrentIndex(1)
        self.combo_a.blockSignals(False)
        self.combo_b.blockSignals(False)
        self._update()

    def _days_changed(self) -> None:
        self._update()

    def _on_combo_changed(self) -> None:
        if len(self._mods) >= 2 and self.combo_a.currentData() == self.combo_b.currentData():
            # auto-move B to the other mod so they can never be identical
            ids = [int(m["id"]) for m in self._mods]
            a = self.combo_a.currentData()
            self.combo_b.blockSignals(True)
            for i, mid in enumerate(ids):
                if mid != a:
                    self.combo_b.setCurrentIndex(i)
                    break
            self.combo_b.blockSignals(False)
        self._update()

    def _update(self) -> None:
        storage = self._storage
        has = bool(self._mods)
        ready = has and len(self._mods) >= 2 and storage is not None
        self.placeholder.setVisible(not ready)
        self.table.setVisible(ready)
        self.grid_host.setVisible(ready)
        for card in (self.plot_cum, self.plot_daily, self.plot_diff):
            card.clear_chart()
        self.table.setRowCount(0)
        if not ready:
            for combo in (self.combo_a, self.combo_b, self.days_combo):
                combo.setEnabled(has)
            self.export_btn.setEnabled(False)
            self.placeholder.setText("Track at least two mods to compare them here." if not has else
                                     "Only one mod tracked. Add a second mod to compare them here.")
            return

        days = int(self.days_combo.currentData() or 60)
        mid_a = int(self.combo_a.currentData())
        mid_b = int(self.combo_b.currentData())
        self._summary_a = analytics.mod_summary(storage, mid_a, days)
        self._summary_b = analytics.mod_summary(storage, mid_b, days)
        a, b = self._summary_a, self._summary_b

        names = {int(m["id"]): m["name"] for m in self._mods}
        self._names = (names.get(mid_a, "A"), names.get(mid_b, "B"))
        name_a, name_b = self._names

        self._fill_table(a, b)
        self._draw_charts(a, b, days)

    # ---- stats table ----------------------------------------------------
    def _fill_table(self, a: Dict[str, Any], b: Dict[str, Any]) -> None:
        def best_day(s: Dict[str, Any]) -> str:
            return f"{s['best_day']['label']} (+{s['best_day']['value']:,})" if s["best_day"] else "—"

        def best_week(s: Dict[str, Any]) -> str:
            return f"{s['best_week']['label']} (+{s['best_week']['value']:,})" if s["best_week"] else "—"

        def milestone_count(s: Dict[str, Any]) -> int:
            return len([m for m in s["milestones"] if m["threshold"] >= 100_000])

        rows: List[Tuple[str, str, str, int, int]] = []  # metric, a_text, b_text, a_win, b_win
        rows.append(("Total downloads", f"{a['total']:,}", f"{b['total']:,}", a["total"], b["total"]))
        rows.append(("Last 7 days", f"+{a['delta_7d']:,}", f"+{b['delta_7d']:,}", a["delta_7d"], b["delta_7d"]))
        rows.append(("Last 30 days", f"+{a['delta_30d']:,}", f"+{b['delta_30d']:,}", a["delta_30d"], b["delta_30d"]))
        rows.append(("Avg / day", f"{a['avg_per_day']}", f"{b['avg_per_day']}", a["avg_per_day"], b["avg_per_day"]))
        rows.append(("Growth (30d)", f"{a['growth_pct']}%", f"{b['growth_pct']}%", a["growth_pct"], b["growth_pct"]))
        rows.append(("Best day", best_day(a), best_day(b), a["best_day"]["value"] if a["best_day"] else 0,
                     b["best_day"]["value"] if b["best_day"] else 0))
        rows.append(("Best week", best_week(a), best_week(b), a["best_week"]["value"] if a["best_week"] else 0,
                     b["best_week"]["value"] if b["best_week"] else 0))
        rows.append(("Next week (est.)", f"{a['next_week_estimate']:,}" if a["next_week_estimate"] else "—",
                     f"{b['next_week_estimate']:,}" if b["next_week_estimate"] else "—",
                     a["next_week_estimate"] or 0, b["next_week_estimate"] or 0))
        rows.append(("Milestones", f"{milestone_count(a)}", f"{milestone_count(b)}",
                     milestone_count(a), milestone_count(b)))
        rows.append(("First seen", str(a["first_seen"] or "—"), str(b["first_seen"] or "—"), 0, 0))

        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for r, (metric, ta, tb, va, vb) in enumerate(rows):
            m_item = QTableWidgetItem(metric)
            m_item.setForeground(QColor(GRAY))
            a_item = QTableWidgetItem(ta)
            b_item = QTableWidgetItem(tb)
            if va > vb:
                a_item.setForeground(QColor(SUCCESS))
            elif va < vb:
                b_item.setForeground(QColor(SUCCESS))
            self.table.setItem(r, 0, m_item)
            self.table.setItem(r, 1, a_item)
            self.table.setItem(r, 2, b_item)
        self.table.resizeColumnsToContents()

    # ---- charts ---------------------------------------------------------
    def _draw_charts(self, a: Dict[str, Any], b: Dict[str, Any], days: int) -> None:
        name_a, name_b = self._names
        aligned = analytics.aligned_totals(a["series"], b["series"], days)

        self.plot_cum.set_ylabel("Total downloads")
        dates = [t[0] for t in aligned]
        self.plot_cum.add_line(dates, [t[1] for t in aligned], self.COLOR_A, width=2, name=name_a)
        self.plot_cum.add_line(dates, [t[2] for t in aligned], self.COLOR_B, width=2, name=name_b)
        self.plot_cum.info.setText(f"{name_a} ({self.COLOR_A})   vs   {name_b} ({self.COLOR_B})")

        self.plot_daily.set_ylabel("Downloads")
        if a["deltas"]:
            self.plot_daily.add_bars(*zip(*a["deltas"]), self.COLOR_A, name=name_a, alpha=110)
        if b["deltas"]:
            self.plot_daily.add_bars(*zip(*b["deltas"]), self.COLOR_B, name=name_b, alpha=110)
        self.plot_daily.info.setText(f"{name_a} (blue)   vs   {name_b} (green)")

        self.plot_diff.set_ylabel("Downloads ahead")
        ahead_dates, ahead_vals = [], []
        behind_dates, behind_vals = [], []
        for day, va, vb in aligned:
            diff = va - vb
            if diff >= 0:
                ahead_dates.append(day)
                ahead_vals.append(diff)
            else:
                behind_dates.append(day)
                behind_vals.append(-diff)
        if ahead_dates:
            self.plot_diff.add_bars(ahead_dates, ahead_vals, SUCCESS, name=f"{name_a} ahead", alpha=170)
        if behind_dates:
            self.plot_diff.add_bars(behind_dates, behind_vals, ERROR, name=f"{name_b} ahead", alpha=170)

    # ---- export ---------------------------------------------------------
    def _export_charts(self) -> None:
        dir_path = QFileDialog.getExistingDirectory(self, "Choose a folder for chart PNGs")
        if not dir_path:
            return
        try:
            written = []
            for slug, card in (("cumulative", self.plot_cum), ("daily", self.plot_daily), ("diff", self.plot_diff)):
                path = Path(dir_path) / f"compare_{slug}.png"
                exporter = ImageExporter(card.plot.getPlotItem())
                exporter.parameters()["width"] = 1200
                exporter.export(str(path))
                written.append(path.name)
            QMessageBox.information(self, "Charts exported", "Saved:\n" + "\n".join(written))
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))


class InsightCard(QFrame):
    """Single insight line with a sentiment-colored accent bar."""

    KIND_ICON = {"positive": "📈", "negative": "⚠️", "info": "💡"}
    KIND_COLOR = {"positive": SUCCESS, "negative": ERROR, "info": ACCENT}

    def __init__(self, insight: Dict[str, Any], parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("Panel")
        kind = insight.get("kind", "info")
        color = self.KIND_COLOR.get(kind, ACCENT)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(12)

        bar = QFrame()
        bar.setFixedWidth(4)
        bar.setStyleSheet(f"background: {color}; border: none; border-radius: 2px;")
        lay.addWidget(bar)

        icon = QLabel(self.KIND_ICON.get(kind, "💡"))
        icon.setStyleSheet("font-size: 20px; border: none; background: transparent;")
        lay.addWidget(icon, 0, Qt.AlignmentFlag.AlignTop)

        col = QVBoxLayout()
        col.setSpacing(2)
        head = QHBoxLayout()
        title = QLabel(insight.get("title", ""))
        title.setStyleSheet(f"font-weight: 700; font-size: 13px; color: {color}; border: none; background: transparent;")
        head.addWidget(title)
        head.addStretch(1)
        if insight.get("mod"):
            mod = QLabel(insight["mod"])
            mod.setStyleSheet(f"color: {GRAY}; font-size: 11px; border: none; background: transparent;")
            head.addWidget(mod)
        col.addLayout(head)
        detail = QLabel(insight.get("detail", ""))
        detail.setWordWrap(True)
        detail.setStyleSheet("font-size: 12px; color: #334155; border: none; background: transparent;")
        col.addWidget(detail)
        lay.addLayout(col, 1)


class InsightsPage(QWidget):
    """Feed of rule-based insight cards generated from recent history."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        header = QHBoxLayout()
        title = QLabel("Insights")
        title.setObjectName("PageTitle")
        hint = QLabel("Automated takeaways from your recent history (plain math, no AI).")
        hint.setObjectName("Hint")
        col = QVBoxLayout()
        col.setSpacing(2)
        col.addWidget(title)
        col.addWidget(hint)
        header.addLayout(col)
        header.addStretch(1)
        layout.addLayout(header)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        host = QWidget()
        self.feed = QVBoxLayout(host)
        self.feed.setContentsMargins(0, 0, 8, 0)
        self.feed.setSpacing(10)
        self.feed.addStretch(1)
        self.scroll.setWidget(host)
        layout.addWidget(self.scroll, 1)

        self.placeholder = QLabel("No insights yet. Poll a few days of data first.")
        self.placeholder.setObjectName("Hint")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)

    def refresh(self, storage: Storage) -> None:
        while self.feed.count():
            item = self.feed.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        insights = analytics.generate_insights(storage, 30)
        if not insights:
            self.feed.addWidget(self.placeholder)
        else:
            for ins in insights:
                self.feed.addWidget(InsightCard(ins))
        self.feed.addStretch(1)


class SearchResultsPage(QWidget):
    """Global search results grouped by category (mods / comments / events / history)."""

    open_url = pyqtSignal(str)

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        self.title = QLabel("Search")
        self.title.setObjectName("PageTitle")
        layout.addWidget(self.title)
        self.hint = QLabel("Type in the search bar to look across mods, comments, events and history.")
        self.hint.setObjectName("Hint")
        layout.addWidget(self.hint)

        self.tree = QTreeWidget()
        self.tree.setColumnCount(3)
        self.tree.setHeaderLabels(["Type", "Match", "Detail"])
        self.tree.setRootIsDecorated(True)
        self.tree.setAlternatingRowColors(True)
        self.tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tree.header().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.tree.itemDoubleClicked.connect(self._open_item)
        layout.addWidget(self.tree, 1)

    def apply_filter(self, text: str) -> None:
        needle = (text or "").strip()
        storage = getattr(self, "_storage", None)
        self.tree.clear()
        if not needle:
            self.hint.setText("Type at least 2 characters in the search bar to search everything.")
            return
        if len(needle) < 2 or storage is None:
            return
        results = storage.search(needle)
        self.hint.setText(f"Results for \u201c{needle}\u201d")
        total = 0
        for group, rows, icon in (
            ("Mods", results["mods"], "📦"),
            ("Comments", results["comments"], "💬"),
            ("Events", results["events"], "🔔"),
            ("History", results["history"], "📈"),
        ):
            if not rows:
                continue
            section = QTreeWidgetItem([f"{icon} {group} ({len(rows)})", "", ""])
            section.setExpanded(True)
            for r in rows:
                child = self._row_item(group, r)
                if child is not None:
                    section.addChild(child)
                    total += 1
            self.tree.addTopLevelItem(section)
        if total == 0:
            self.tree.addTopLevelItem(QTreeWidgetItem([f"No matches for \u201c{needle}\u201d.", "", ""]))
        self.tree.expandAll()

    def _row_item(self, group: str, row: Dict[str, Any]) -> Optional[QTreeWidgetItem]:
        url = str(row.get("url") or row.get("mod_url") or "")
        if group == "Mods":
            item = QTreeWidgetItem(["Mod", row["name"], row.get("content_type", "mod")])
        elif group == "Comments":
            content = " ".join(str(row.get("content") or "").split())
            item = QTreeWidgetItem(["Comment", f"{row['author']}: {content[:120]}", f"{row.get('mod_name', '')} · {row.get('posted_at', '')}"])
        elif group == "Events":
            item = QTreeWidgetItem(["Event", str(row.get("message", ""))[:160], f"{row.get('mod_name') or ''} · {row.get('created_at', '')}"])
        elif group == "History":
            item = QTreeWidgetItem(["History", f"{row.get('mod_name', '')} — {int(row.get('downloads_total') or 0):,} downloads", row.get("fetched_at", "")])
        else:
            return None
        if url:
            item.setData(0, Qt.ItemDataRole.UserRole, url)
        return item

    def _open_item(self, item: QTreeWidgetItem, _column: int) -> None:
        url = item.data(0, Qt.ItemDataRole.UserRole)
        if url:
            self.open_url.emit(str(url))


class BadgeCard(QFrame):
    """Single achievement badge, highlighted when unlocked."""

    ICONS = {
        "tracked": "🛰️",
        "milestone-100000": "🚩",
        "milestone-250000": "🥉",
        "milestone-500000": "🥈",
        "milestone-1000000": "🥇",
        "milestone-2500000": "💎",
        "milestone-5000000": "👑",
        "milestone-10000000": "🚀",
        "best-week": "🏆",
        "steady": "📈",
        "big-day": "🌋",
        "community": "💬",
        "fast-riser": "⚡",
    }

    def __init__(self, badge: Dict[str, Any], parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("Panel")
        unlocked = bool(badge.get("unlocked"))
        self.setStyleSheet(
            f"QFrame#Panel {{ border: 1px solid {ACCENT}; }}" if unlocked
            else f"QFrame#Panel {{ border: 1px solid {BORDER}; }}"
        )
        icon = self.ICONS.get(badge.get("key", ""), "🎖️") if unlocked else "🔒"
        border_color = ACCENT if unlocked else FAINT
        text_color = TEXT if unlocked else FAINT

        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 12)
        lay.setSpacing(4)
        top = QHBoxLayout()
        ic = QLabel(icon)
        ic.setStyleSheet(f"font-size: 22px; background: transparent; border: none; color: {border_color};")
        top.addWidget(ic)
        top.addStretch(1)
        if unlocked and badge.get("date"):
            d = QLabel(str(badge["date"]))
            d.setStyleSheet(f"color: {GRAY}; font-size: 10px; background: transparent; border: none;")
            top.addWidget(d)
        lay.addLayout(top)
        title = QLabel(badge.get("title", ""))
        title.setWordWrap(True)
        title.setStyleSheet(f"font-weight: 700; font-size: 13px; color: {text_color}; background: transparent; border: none;")
        lay.addWidget(title)
        detail = QLabel(badge.get("detail", ""))
        detail.setWordWrap(True)
        detail.setStyleSheet(f"color: {GRAY}; font-size: 11px; background: transparent; border: none;")
        lay.addWidget(detail)


class AchievementsPage(QWidget):
    """Milestone timeline and achievement badges for each tracked mod."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._storage: Optional[Storage] = None
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(0, 0, 8, 0)
        layout.setSpacing(12)

        title = QLabel("Achievements")
        title.setObjectName("PageTitle")
        layout.addWidget(title)
        hint = QLabel("Milestone timeline and achievement badges for each tracked mod.")
        hint.setObjectName("Hint")
        layout.addWidget(hint)

        row = QHBoxLayout()
        row.setSpacing(10)
        row.addWidget(QLabel("Mod:"))
        self.mod_combo = QComboBox()
        self.mod_combo.setMinimumWidth(240)
        self.mod_combo.currentIndexChanged.connect(self._render)
        row.addWidget(self.mod_combo)
        row.addStretch(1)
        layout.addLayout(row)

        self.summary = QFrame()
        self.summary.setObjectName("Panel")
        sl = QHBoxLayout(self.summary)
        sl.setContentsMargins(14, 10, 14, 10)
        sl.setSpacing(24)
        self.sum_labels: Dict[str, QLabel] = {}
        for key, cap in (("total", "Total downloads"), ("first", "First seen"), ("span", "Days tracked"), ("avg", "Avg / day")):
            col = QVBoxLayout()
            col.setSpacing(2)
            c = QLabel(cap)
            c.setObjectName("Caption")
            v = QLabel("—")
            v.setStyleSheet("font-weight: 700; font-size: 16px; background: transparent; border: none;")
            self.sum_labels[key] = v
            col.addWidget(c)
            col.addWidget(v)
            sl.addLayout(col)
        sl.addStretch(1)
        layout.addWidget(self.summary)

        layout.addWidget(section_label("Milestones"))
        self.timeline = QFrame()
        self.timeline.setObjectName("Panel")
        self.tl_lay = QVBoxLayout(self.timeline)
        self.tl_lay.setContentsMargins(14, 12, 14, 12)
        self.tl_lay.setSpacing(8)
        layout.addWidget(self.timeline)

        layout.addWidget(section_label("Achievements"))
        self.badges_host = QWidget()
        self.badges_flow = FlowLayout(self.badges_host, spacing=10, min_width=220)
        layout.addWidget(self.badges_host)

        self.placeholder = QLabel("No tracked mods yet. Poll or add mods to see achievements.")
        self.placeholder.setObjectName("Hint")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setMinimumHeight(120)
        layout.addWidget(self.placeholder, 1)

        self.scroll.setWidget(content)
        outer.addWidget(self.scroll)

    def refresh(self, storage: Storage) -> None:
        self._storage = storage
        current = self.mod_combo.currentData()
        self.mod_combo.blockSignals(True)
        self.mod_combo.clear()
        for m in storage.get_mods(active_only=True):
            self.mod_combo.addItem(m["name"], int(m["id"]))
        if current is not None:
            idx = self.mod_combo.findData(current)
            if idx >= 0:
                self.mod_combo.setCurrentIndex(idx)
        self.mod_combo.blockSignals(False)
        self._render()

    def _render(self) -> None:
        storage = self._storage
        has = storage is not None and self.mod_combo.count() > 0
        self.summary.setVisible(has)
        self.timeline.setVisible(has)
        self.badges_host.setVisible(has)
        self.placeholder.setVisible(not has)
        if not has:
            return
        data = analytics.achievements(storage, int(self.mod_combo.currentData()))
        tl = data["milestones"]
        self.sum_labels["total"].setText(f"{tl['total']:,}")
        self.sum_labels["first"].setText(str(tl["first_seen"] or "—"))
        span = (datetime.date.today() - tl["first_seen"]).days + 1 if tl["first_seen"] else 0
        self.sum_labels["span"].setText(f"{span:,}" if span else "—")
        self.sum_labels["avg"].setText(f"{tl['avg_per_day']:,}")

        while self.tl_lay.count():
            item = self.tl_lay.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        for m in tl["reached"]:
            self.tl_lay.addWidget(self._tl_row(
                f"◆ {m['threshold']:,} downloads",
                f"Reached {m['date']} — {m['days']} days after the previous one ({m['per_day']} per day).",
                SUCCESS))
        if tl["next"]:
            n = tl["next"]
            eta = f"~{n['eta_days']} days at {n['avg_per_day']}/day" if n["eta_days"] else "pace too slow to estimate"
            self.tl_lay.addWidget(self._tl_row(
                f"◇ {n['threshold']:,} downloads",
                f"{n['total']:,} so far — {n['remaining']:,} to go ({eta}).",
                WARNING, progress=n["total"] / n["threshold"]))
        if not tl["reached"] and not tl["next"]:
            self.tl_lay.addWidget(self._tl_row("No milestones yet", "Keep polling — milestones appear as downloads grow.", GRAY))
        self.tl_lay.addStretch(1)

        self.badges_flow.clear()
        for b in data["achievements"]:
            self.badges_flow.add_card(BadgeCard(b))

    @staticmethod
    def _tl_row(title: str, detail: str, color: str, progress: Optional[float] = None) -> QFrame:
        row = QFrame()
        row.setObjectName("EventRow")
        h = QHBoxLayout(row)
        h.setContentsMargins(4, 4, 4, 4)
        h.setSpacing(12)
        dot = QLabel()
        dot.setFixedSize(12, 12)
        dot.setStyleSheet(f"background: {color}; border-radius: 6px; border: none;")
        h.addWidget(dot)
        col = QVBoxLayout()
        col.setSpacing(1)
        t = QLabel(title)
        t.setStyleSheet(f"font-weight: 600; font-size: 13px; color: {color}; background: transparent; border: none;")
        col.addWidget(t)
        d = QLabel(detail)
        d.setWordWrap(True)
        d.setStyleSheet(f"color: {GRAY}; font-size: 11px; background: transparent; border: none;")
        col.addWidget(d)
        if progress is not None:
            bar = QProgressBar()
            bar.setRange(0, 1000)
            bar.setValue(int(min(1.0, progress) * 1000))
            bar.setTextVisible(False)
            bar.setFixedHeight(6)
            col.addWidget(bar)
        h.addLayout(col, 1)
        return row


class ActivityFeed(QFrame):
    """Sidebar panel listing recent download / comment events."""

    KIND_STYLE = {
        "download": ("⬇", SUCCESS),
        "comment": ("💬", ACCENT),
        "reply": ("↩", WARNING),
    }

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("Panel")
        header = QLabel("ACTIVITY")
        header.setObjectName("SectionTitle")
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self._container = QWidget()
        self._container.setStyleSheet("background: transparent;")
        self._list_lay = QVBoxLayout(self._container)
        self._list_lay.setContentsMargins(2, 2, 2, 2)
        self._list_lay.setSpacing(2)
        self._list_lay.addStretch(1)
        self._scroll.setWidget(self._container)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 14, 14, 12)
        lay.setSpacing(10)
        lay.addWidget(header)
        lay.addWidget(self._scroll)
        self._empty = None

    def refresh(self, storage: Storage) -> None:
        while self._list_lay.count() > 1:
            item = self._list_lay.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        if self._empty:
            self._empty.deleteLater()
            self._empty = None

        rows = storage.recent_events(40)
        if not rows:
            self._empty = QLabel("No activity yet.\n\nRun a poll to start\nrecording events.")
            self._empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._empty.setObjectName("Hint")
            self._empty.setStyleSheet(f"color: {FAINT}; padding: 24px;")
            self._list_lay.insertWidget(0, self._empty)
            return

        for event in rows:
            kind = event["kind"] or "download"
            icon, color = self.KIND_STYLE.get(kind, ("•", GRAY))
            row = QFrame()
            row.setObjectName("EventRow")
            h = QHBoxLayout(row)
            h.setContentsMargins(8, 6, 8, 6)
            h.setSpacing(8)

            icon_label = QLabel(icon)
            icon_label.setStyleSheet(f"color: {color}; font-size: 14px; background: transparent;")
            icon_label.setFixedWidth(20)

            body = QVBoxLayout()
            body.setSpacing(1)
            head = QLabel(str(event["mod_name"] or "ModDB"))
            head.setStyleSheet(f"color: {color}; font-weight: 700; font-size: 12px; background: transparent;")
            msg = QLabel(str(event["message"]))
            msg.setWordWrap(True)
            msg.setStyleSheet(f"color: {GRAY}; font-size: 12px; background: transparent;")
            body.addWidget(head)
            body.addWidget(msg)

            stamp = QLabel(relative_time(event["created_at"]))
            stamp.setStyleSheet(f"color: {FAINT}; font-size: 11px; background: transparent;")
            stamp.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignRight)

            h.addWidget(icon_label)
            h.addLayout(body, 1)
            h.addWidget(stamp)
            self._list_lay.insertWidget(self._list_lay.count() - 1, row)


class ModCard(QFrame):
    """Compact card for a single tracked mod on the My Mods page."""

    open_requested = pyqtSignal(str)
    refresh_requested = pyqtSignal(int)
    favorite_toggled = pyqtSignal(str, bool)
    remove_requested = pyqtSignal(str)
    export_requested = pyqtSignal()

    TYPE_BADGE = {
        "mod": ("MOD", ACCENT),
        "addon": ("ADDON", SUCCESS),
        "file": ("FILE", WARNING),
    }

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("ModCard")
        self.setFixedHeight(148)
        self._url = ""
        self._name_id = ""
        self._mod_id = 0
        self._favorite = False

        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_menu)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(6)

        # header: badge + name + favorite star
        self._badge = QLabel("MOD")
        self._badge.setFixedSize(54, 20)
        self._badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._badge.setStyleSheet(
            "font-size: 9px; font-weight: 800; letter-spacing: 1px; border-radius: 4px;"
        )
        self._name = QLabel()
        self._name.setStyleSheet("font-size: 14px; font-weight: 700;")
        self._name.setToolTip("")
        self._full_name = ""
        self._star = QToolButton()
        self._star.setFixedSize(26, 26)
        self._star.setCursor(Qt.CursorShape.PointingHandCursor)
        self._star.setToolTip("Toggle favorite")
        self._star.clicked.connect(self._toggle_favorite)
        header = QHBoxLayout()
        header.setSpacing(8)
        header.addWidget(self._badge)
        header.addWidget(self._name, 1)
        header.addWidget(self._star)
        lay.addLayout(header)

        # numbers
        self._dl = QLabel("—")
        self._dl.setObjectName("StatValue")
        self._dl.setStyleSheet("font-size: 20px;")
        self._dl_cap = QLabel("DOWNLOADS")
        self._dl_cap.setObjectName("StatCaption")
        num_col = QVBoxLayout()
        num_col.setSpacing(0)
        num_col.addWidget(self._dl)
        num_col.addWidget(self._dl_cap)
        self._growth = QLabel("")
        self._growth.setStyleSheet(f"color: {SUCCESS}; font-size: 12px; font-weight: 600;")
        self._meta = QLabel("")
        self._meta.setStyleSheet(f"color: {GRAY}; font-size: 12px;")
        right_col = QVBoxLayout()
        right_col.setSpacing(3)
        right_col.addStretch(1)
        right_col.addWidget(self._growth)
        right_col.addWidget(self._meta)
        numbers = QHBoxLayout()
        numbers.setSpacing(20)
        numbers.addLayout(num_col)
        numbers.addLayout(right_col, 1)
        lay.addLayout(numbers)

        # footer: updated stamp + actions
        self._stamp = QLabel("")
        self._stamp.setStyleSheet(f"color: {FAINT}; font-size: 11px;")
        self._open_btn = QPushButton("Open")
        self._open_btn.clicked.connect(lambda: self.open_requested.emit(self._url))
        self._refresh_btn = QPushButton("Refresh")
        self._refresh_btn.clicked.connect(lambda: self.refresh_requested.emit(self._mod_id))
        footer = QHBoxLayout()
        footer.setSpacing(8)
        footer.addWidget(self._stamp)
        footer.addStretch(1)
        footer.addWidget(self._open_btn)
        footer.addWidget(self._refresh_btn)
        lay.addLayout(footer)

    # ---- data -----------------------------------------------------------
    def set_data(self, totals: Dict[str, Any], meta: Dict[str, Any],
                 delta_7d: int = 0, updated: str = "") -> None:
        self._mod_id = int(totals["id"])
        self._name_id = str(totals.get("name_id") or meta.get("name_id") or "")
        self._url = str(totals.get("url") or meta.get("url") or "")
        self._favorite = bool(totals.get("favorite", meta.get("favorite", False)))
        content_type = str(meta.get("content_type") or totals.get("content_type") or "mod")
        name = str(totals.get("name") or meta.get("name") or self._name_id)

        badge_text, badge_color = self.TYPE_BADGE.get(content_type, ("MOD", ACCENT))
        self._badge.setText(badge_text)
        self._badge.setStyleSheet(
            "font-size: 9px; font-weight: 800; letter-spacing: 1px; border-radius: 4px;"
            f" background: {badge_color}22; color: {badge_color}; border: 1px solid {badge_color}55;"
        )
        self._name.setText(name)
        self._full_name = name
        self._name.setToolTip(name)
        self._refresh_name_elide()

        self._dl.setText(format_num(totals.get("downloads_total")))
        self._dl_cap.setText(f"DOWNLOADS · +{format_num(totals.get('downloads_today'))} TODAY")
        if delta_7d > 0:
            self._growth.setText(f"▲ +{delta_7d:,} last 7 days")
            self._growth.setStyleSheet(f"color: {SUCCESS}; font-size: 12px; font-weight: 600;")
        else:
            self._growth.setText("— no change this week")
            self._growth.setStyleSheet(f"color: {GRAY}; font-size: 12px;")
        rank = totals.get("rank")
        parts = []
        if rank is not None:
            parts.append(f"Rank {rank}{'/' + str(totals['rank_total']) if totals.get('rank_total') else ''}")
        parts.append(f"Watchers {format_num(totals.get('watchers'))}")
        parts.append(f"Visits {format_num(totals.get('visits'))}")
        parts.append(f"Comments {totals.get('comments', 0)}")
        parts.append(f"Replies {totals.get('replies', 0)}")
        self._meta.setText(" · ".join(parts))
        self._stamp.setText(f"Updated {relative_time(updated)}" if updated else "")

        star = "★" if self._favorite else "☆"
        self._star.setText(star)
        color = WARNING if self._favorite else FAINT
        hover = f" QToolButton:hover {{ color: {WARNING}; }}" if not self._favorite else ""
        self._star.setStyleSheet(
            "QToolButton { border: none; background: transparent; font-size: 20px;"
            f" color: {color}; }}" + hover
        )

    def matches(self, text: str) -> bool:
        needle = text.strip().lower()
        if not needle:
            return True
        hay = f"{self._name.text()} {self._name_id} {self._url}".lower()
        return needle in hay

    # ---- actions --------------------------------------------------------
    def _refresh_name_elide(self) -> None:
        if not self._full_name:
            return
        fm = self._name.fontMetrics()
        available = max(60, self._name.width())
        self._name.setText(fm.elidedText(self._full_name, Qt.TextElideMode.ElideRight, available))

    def resizeEvent(self, event) -> None:  # noqa: N802
        super().resizeEvent(event)
        self._refresh_name_elide()

    def _toggle_favorite(self) -> None:
        self._favorite = not self._favorite
        self._star.setText("★" if self._favorite else "☆")
        self._star.setStyleSheet(
            "QToolButton { border: none; background: transparent; font-size: 20px;"
            f" color: {WARNING if self._favorite else FAINT}; }}"
        )
        self.favorite_toggled.emit(self._name_id, self._favorite)

    def _show_menu(self, pos) -> None:
        menu = QMenu(self)
        fav_action = menu.addAction("Unfavorite" if self._favorite else "Favorite")
        menu.addSeparator()
        menu.addAction("Refresh mod...")
        open_action = menu.addAction("Open on ModDB")
        copy_action = menu.addAction("Copy URL")
        menu.addSeparator()
        export_action = menu.addAction("Export JSON...")
        remove_action = menu.addAction("Remove from tracking")
        chosen = menu.exec(self.mapToGlobal(pos))
        if chosen == fav_action:
            self._toggle_favorite()
        elif chosen == open_action:
            self.open_requested.emit(self._url)
        elif chosen == copy_action:
            QApplication.clipboard().setText(self._url)
        elif chosen == export_action:
            self.export_requested.emit()
        elif chosen == remove_action:
            self.remove_requested.emit(self._name_id)


# --------------------------------------------------------------------------
# pages
# --------------------------------------------------------------------------

class DashboardPage(QWidget):
    regen_requested = pyqtSignal()
    view_insights = pyqtSignal()
    SKIP_CHARTS = {"dashboard.png"}
    CHART_ORDER = [
        "downloads_per_day.png",
        "total_downloads.png",
        "mod_overview.png",
        "comment_activity.png",
    ]
    MIN_COLUMN_WIDTH = 460

    def __init__(self, config: Dict[str, Any], config_path: str = CONFIG_FILE, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._config = config
        self._config_path = config_path
        self._storage: Optional[Storage] = None
        self._insights_available = False
        self._chart_cards: List[ChartCard] = []
        self._current_cols = -1
        self._load_prefs()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        header = QHBoxLayout()
        self.title = QLabel("Dashboard")
        self.title.setObjectName("PageTitle")
        self.subtitle = QLabel("Never synced")
        self.subtitle.setObjectName("Hint")
        head_col = QVBoxLayout()
        head_col.setSpacing(2)
        head_col.addWidget(self.title)
        head_col.addWidget(self.subtitle)
        header.addLayout(head_col)
        header.addStretch(1)
        self.customize_btn = QToolButton()
        self.customize_btn.setText("Customize \u25be")
        self.customize_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.customize_btn.setToolTip("Rearrange dashboard widgets")
        self.customize_btn.setMenu(self._build_customize_menu())
        header.addWidget(self.customize_btn)
        layout.addLayout(header)

        self.stats_host = QWidget()
        self.stat_grid = QGridLayout(self.stats_host)
        self.stat_grid.setSpacing(12)
        self.stat_cards: Dict[str, StatCard] = {}
        for key, caption in [
            ("total", "Total downloads"),
            ("today", "Downloads today"),
            ("week", "This week"),
            ("month", "This month"),
            ("avg", "Average / day"),
            ("comments", "Comments"),
            ("replies", "Replies"),
            ("fastest", "Fastest growing"),
        ]:
            card = StatCard(caption)
            self.stat_cards[key] = card
        for i, key in enumerate(["total", "today", "week", "month"]):
            self.stat_grid.addWidget(self.stat_cards[key], 0, i)
        for i, key in enumerate(["avg", "comments", "replies", "fastest"]):
            self.stat_grid.addWidget(self.stat_cards[key], 1, i)
        layout.addWidget(self.stats_host)

        self._build_insights_strip()
        layout.addWidget(self.insights_panel)

        self.charts_section = self._build_charts_section()
        self.body = QHBoxLayout()
        self.body.setSpacing(12)
        self.body.addWidget(self.charts_section, 1)
        self.activity = ActivityFeed()
        self.activity.setFixedWidth(320)
        self.body.addWidget(self.activity)
        layout.addLayout(self.body, 1)

        self.below_row = QHBoxLayout()
        self.below_row.setSpacing(12)
        layout.addLayout(self.below_row)

        self._apply_layout()

    # ---- rearrangement --------------------------------------------------
    def _build_customize_menu(self) -> QMenu:
        menu = QMenu(self)
        self._act_stats = QAction("Stat cards", self)
        self._act_insights = QAction("Insights", self)
        self._act_charts = QAction("Charts", self)
        self._act_activity = QAction("Activity", self)
        for act, key in ((self._act_stats, "stats"), (self._act_insights, "insights"),
                         (self._act_charts, "charts"), (self._act_activity, "activity")):
            act.setCheckable(True)
            act.toggled.connect(lambda on, k=key: self._set_pref(k, on))
            menu.addAction(act)
        menu.addSeparator()
        sub = menu.addMenu("Activity position")
        self._act_pos_right = QAction("Right side", self)
        self._act_pos_below = QAction("Below charts", self)
        for act, pos in ((self._act_pos_right, "right"), (self._act_pos_below, "below")):
            act.setCheckable(True)
            act.triggered.connect(lambda _, p=pos: self._set_activity_position(p))
            sub.addAction(act)
        return menu

    def _load_prefs(self) -> None:
        prefs = self._config.get("ui", {}).get("dashboard", {}) or {}
        self._prefs = {
            "stats": bool(prefs.get("stats", True)),
            "insights": bool(prefs.get("insights", True)),
            "charts": bool(prefs.get("charts", True)),
            "activity": bool(prefs.get("activity", True)),
            "activity_position": "below" if prefs.get("activity_position") == "below" else "right",
        }

    def _save_prefs(self) -> None:
        try:
            self._config.setdefault("ui", {}).setdefault("dashboard", {}).update(self._prefs)
            save_config(self._config, self._config_path)
        except Exception:  # noqa: BLE001
            pass

    def _sync_menu(self) -> None:
        for act, key in ((self._act_stats, "stats"), (self._act_insights, "insights"),
                         (self._act_charts, "charts"), (self._act_activity, "activity")):
            act.setChecked(self._prefs[key])
        self._act_pos_right.setChecked(self._prefs["activity_position"] == "right")
        self._act_pos_below.setChecked(self._prefs["activity_position"] == "below")

    def _set_pref(self, key: str, value: bool) -> None:
        if self._prefs.get(key) == bool(value):
            return
        self._prefs[key] = bool(value)
        self._apply_layout()
        self._save_prefs()

    def _set_activity_position(self, pos: str) -> None:
        if self._prefs["activity_position"] == pos:
            return
        self._prefs["activity_position"] = pos
        self._move_activity()
        self._sync_menu()
        self._save_prefs()

    def _move_activity(self) -> None:
        pos = self._prefs["activity_position"]
        is_below = self.below_row.indexOf(self.activity) >= 0
        if pos == "below":
            if not is_below:
                self.body.removeWidget(self.activity)
                self.below_row.addWidget(self.activity, 1)
            self.activity.setFixedWidth(0)
            self.activity.setMaximumHeight(230)
        else:
            if is_below:
                self.below_row.removeWidget(self.activity)
                self.body.addWidget(self.activity)
            self.activity.setFixedWidth(320)
            self.activity.setMaximumHeight(16777215)

    def _apply_layout(self) -> None:
        self.stats_host.setVisible(self._prefs["stats"])
        self.charts_section.setVisible(self._prefs["charts"])
        self.activity.setVisible(self._prefs["activity"])
        self.insights_panel.setVisible(self._prefs["insights"] and self._insights_available)
        self._move_activity()
        self._sync_menu()

    def _build_insights_strip(self) -> None:
        self.insights_panel = QFrame()
        self.insights_panel.setObjectName("Panel")
        ins = QHBoxLayout(self.insights_panel)
        ins.setContentsMargins(14, 10, 14, 10)
        ins.setSpacing(16)
        label = QLabel("INSIGHTS")
        label.setObjectName("SectionTitle")
        ins.addWidget(label)
        self.insight_items: List[QLabel] = []
        for _ in range(3):
            lab = QLabel("—")
            lab.setWordWrap(True)
            lab.setStyleSheet("font-size: 12px; color: #334155; background: transparent; border: none;")
            self.insight_items.append(lab)
            ins.addWidget(lab, 1)
        view_all = QPushButton("View all")
        view_all.setCursor(Qt.CursorShape.PointingHandCursor)
        view_all.clicked.connect(self._request_insights)
        ins.addWidget(view_all)
        self.insights_panel.setVisible(False)

    def _build_charts_section(self) -> QWidget:
        wrap = QWidget()
        wl = QVBoxLayout(wrap)
        wl.setContentsMargins(0, 0, 0, 0)
        wl.setSpacing(8)
        chart_header = QHBoxLayout()
        chart_header.addWidget(section_label("Charts"))
        chart_header.addStretch(1)
        self.regen_btn = QPushButton("Regenerate charts")
        self.regen_btn.clicked.connect(self._request_regen)
        chart_header.addWidget(self.regen_btn)
        wl.addLayout(chart_header)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self.charts_host = QWidget()
        self.charts_grid = QGridLayout(self.charts_host)
        self.charts_grid.setSpacing(16)
        self.charts_grid.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.scroll.setWidget(self.charts_host)
        wl.addWidget(self.scroll, 1)
        self.placeholder = QLabel("No charts yet. Click \u201cPoll now\u201d to fetch data.")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setObjectName("Hint")
        self.placeholder.setMinimumHeight(180)
        return wrap

    def refresh(self, storage: Optional[Storage] = None) -> None:
        if storage is not None:
            self._storage = storage
            self._update_stats(storage)
            self._update_insights(storage)
            self.activity.refresh(storage)
        self._reload_charts()

    def _request_regen(self) -> None:
        self.regen_requested.emit()

    def _request_insights(self) -> None:
        self.view_insights.emit()

    def _update_insights(self, storage: Storage) -> None:
        insights = analytics.generate_insights(storage, 30, limit=3)
        self._insights_available = bool(insights)
        self.insights_panel.setVisible(self._prefs["insights"] and self._insights_available)
        for i, lab in enumerate(self.insight_items):
            if i < len(insights):
                lab.setText(f"🛈 {insights[i]['title']} — {insights[i]['detail']}")
            else:
                lab.setText("")

    def reload(self) -> None:
        self._reload_charts()

    def _reload_charts(self) -> None:
        while self.charts_grid.count():
            item = self.charts_grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._chart_cards = []
        self._current_cols = -1

        out_dir = Path(self._config["paths"]["output"])
        available = {p.name: p for p in out_dir.glob("*.png") if p.name not in self.SKIP_CHARTS}
        ordered = [available[name] for name in self.CHART_ORDER if name in available]
        ordered += sorted(p for name, p in available.items() if name not in self.CHART_ORDER)

        if not ordered:
            self.charts_grid.addWidget(self.placeholder, 0, 0)
            return

        for png in ordered:
            card = ChartCard()
            card.set_image(png)
            self._chart_cards.append(card)
        self._relayout()

    def _update_stats(self, storage: Storage) -> None:
        s = storage.dashboard_stats()
        today = int(s["today_downloads"] or 0)
        self.stat_cards["total"].set_value(s["total_downloads"], f"▲ +{today:,} today", SUCCESS)
        self.stat_cards["today"].set_value(today, "across all tracked items", GRAY)
        self.stat_cards["week"].set_value(s["week_downloads"], "last 7 days", GRAY)
        self.stat_cards["month"].set_value(s["month_downloads"], f"{s['avg_per_day']} per day", GRAY)
        self.stat_cards["avg"].set_value(s["avg_per_day"], "30-day average", GRAY)
        self.stat_cards["comments"].set_value(s["comments"], f"+{s['replies']} replies", GRAY)
        self.stat_cards["replies"].set_value(s["replies"], "", GRAY)
        if s["fastest_mod"]:
            self.stat_cards["fastest"].set_text(s["fastest_mod"], f"▲ +{s['fastest_delta']:,} this week", SUCCESS)
        else:
            self.stat_cards["fastest"].set_text("—", "no growth yet", GRAY)

        member = storage.meta_get("member_name") or "not set"
        last = storage.meta_get("last_poll")
        self.subtitle.setText(f"Member: {member}   |   Last poll: {last if last else 'never'}")

    def _columns(self) -> int:
        width = self.scroll.viewport().width()
        return 2 if width >= self.MIN_COLUMN_WIDTH * 2 else 1

    def _relayout(self) -> None:
        for i in reversed(range(self.charts_grid.count())):
            self.charts_grid.takeAt(i)
        cols = self._columns()
        self._current_cols = cols
        for i, card in enumerate(self._chart_cards):
            r, c = divmod(i, cols)
            self.charts_grid.addWidget(card, r, c)
        self._rescale()

    def _rescale(self) -> None:
        if not self._chart_cards:
            return
        cols = max(1, self._current_cols)
        width = self.scroll.viewport().width()
        col_width = max(200, (width - (cols - 1) * 16) // cols)
        for card in self._chart_cards:
            card.rescale(col_width)

    def resizeEvent(self, event) -> None:  # noqa: N802
        super().resizeEvent(event)
        if not self._chart_cards:
            return
        if self._columns() != self._current_cols:
            self._relayout()
        else:
            self._rescale()


class ModsPage(QWidget):
    open_requested = pyqtSignal(str)
    refresh_requested = pyqtSignal(int)
    favorite_toggled = pyqtSignal(str, bool)
    remove_requested = pyqtSignal(str)
    export_requested = pyqtSignal()

    def __init__(self, config: Dict[str, Any], config_path: str = CONFIG_FILE, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._config = config
        self._config_path = config_path
        self._cards: List[ModCard] = []
        self._filter = str(config.get("ui", {}).get("mods_filter", ""))
        self._sort = int(config.get("ui", {}).get("mods_sort", 0))
        self._prefs_timer = QTimer(self)
        self._prefs_timer.setSingleShot(True)
        self._prefs_timer.setInterval(400)
        self._prefs_timer.timeout.connect(self._save_ui_prefs)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        header = QHBoxLayout()
        head_col = QVBoxLayout()
        head_col.setSpacing(2)
        self.title = QLabel("My Mods")
        self.title.setObjectName("PageTitle")
        self.count_label = QLabel("")
        self.count_label.setObjectName("Hint")
        head_col.addWidget(self.title)
        head_col.addWidget(self.count_label)
        header.addLayout(head_col)
        header.addStretch(1)
        layout.addLayout(header)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)
        self.search = QLineEdit()
        self.search.setObjectName("Search")
        self.search.setPlaceholderText("Search mods\u2026")
        self.search.setClearButtonEnabled(True)
        self.search.setFixedWidth(260)
        self.search.textChanged.connect(self.apply_filter)
        if self._filter:
            self.search.setText(self._filter)
        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["Sort: Downloads", "Sort: Today", "Sort: 7-day growth", "Sort: Name", "Sort: Favorites first"])
        self.sort_combo.setCurrentIndex(self._sort)
        self.sort_combo.currentIndexChanged.connect(self._sort_changed)
        self.export_btn = QPushButton("Export JSON\u2026")
        self.export_btn.clicked.connect(self.export_requested.emit)
        toolbar.addWidget(self.search)
        toolbar.addWidget(self.sort_combo)
        toolbar.addStretch(1)
        toolbar.addWidget(self.export_btn)
        layout.addLayout(toolbar)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self.host = QWidget()
        self.host.setStyleSheet("background: transparent;")
        self.flow = FlowLayout(self.host, spacing=12, min_width=320)
        self.scroll.setWidget(self.host)
        layout.addWidget(self.scroll, 1)

        self.placeholder = QLabel(
            "Nothing tracked yet.\n\nAdd mods in Configuration \u2192 Manual mods, or click \u201cRescan profile\u201d "
            "to auto-discover your content."
        )
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setObjectName("Hint")
        self.placeholder.setMinimumHeight(220)
        layout.addWidget(self.placeholder)

    def _sort_changed(self, index: int) -> None:
        self._sort = index
        self._rebuild()
        self._save_ui_prefs()

    def _save_ui_prefs(self) -> None:
        try:
            self._config.setdefault("ui", {})["mods_sort"] = self._sort
            self._config["ui"]["mods_filter"] = self._filter
            save_config(self._config, self._config_path)
        except Exception:  # noqa: BLE001
            pass

    def refresh(self, storage: Storage) -> None:
        self._rebuild(storage)

    def _rebuild(self, storage: Optional[Storage] = None) -> None:
        if storage is not None:
            self._storage = storage
        storage = getattr(self, "_storage", None)
        if storage is None:
            return

        totals = storage.totals_per_mod()
        mods = {int(m["id"]): m for m in storage.get_mods(active_only=True)}

        # per-mod deltas (7-day growth) from snapshots
        deltas: Dict[int, int] = {}
        updated_at: Dict[int, str] = {}
        snaps = storage.snapshots_for_all([int(t["id"]) for t in totals])
        today = datetime.date.today()
        for mod_id, rows in snaps.items():
            daily_max: Dict[datetime.date, int] = {}
            for s in rows:
                try:
                    d = datetime.date.fromisoformat(str(s["fetched_at"])[:10])
                except Exception:  # noqa: BLE001
                    continue
                daily_max[d] = max(daily_max.get(d, 0), int(s["downloads_total"] or 0))
            days = sorted(daily_max)
            delta = 0
            prev = None
            for d in days:
                cur = daily_max[d]
                if prev is not None and (today - d).days < 7:
                    delta += max(0, cur - prev)
                prev = cur
            deltas[mod_id] = delta
            if rows:
                updated_at[mod_id] = str(rows[-1]["fetched_at"])

        self.flow.clear()
        self._cards = []
        for t in totals:
            meta = dict(mods.get(int(t["id"]), {}))
            card = ModCard()
            card.set_data(t, meta, delta_7d=deltas.get(int(t["id"]), 0),
                          updated=updated_at.get(int(t["id"]), ""))
            card.open_requested.connect(self.open_requested)
            card.refresh_requested.connect(self.refresh_requested)
            card.favorite_toggled.connect(self.favorite_toggled)
            card.remove_requested.connect(self.remove_requested)
            card.export_requested.connect(self.export_requested)
            self._cards.append(card)
            self.flow.add_card(card)

        self._apply_sort()
        self._apply_filter()
        self.count_label.setText(f"{len(self._cards)} tracked" + ("  ·  matches filter" if self._filter else ""))

        has_any = bool(self._cards) or bool(storage.get_mods(active_only=True))
        self.scroll.setVisible(has_any)
        self.placeholder.setVisible(not has_any)

    def _apply_sort(self) -> None:
        if self._sort == 0:
            key = _card_downloads
        elif self._sort == 1:
            key = _card_today
        elif self._sort == 2:
            key = _card_growth
        elif self._sort == 3:
            key = lambda c: c._name.text().lower()
        else:
            key = lambda c: (not c._favorite, _card_downloads(c))
        self._cards.sort(key=key, reverse=(self._sort in (0, 1, 2, 4)))
        self.flow.clear()
        for card in self._cards:
            self.flow.add_card(card)

    def apply_filter(self, text: str) -> None:
        self._filter = text
        self._apply_filter()
        self._prefs_timer.start()
        if self.count_label.text():
            self.count_label.setText(f"{len(self._cards)} tracked" + ("  ·  matches filter" if text.strip() else ""))

    def _apply_filter(self) -> None:
        needle = self._filter.strip().lower()
        for card in self._cards:
            match = (not needle) or card.matches(self._filter)
            card.setVisible(match)

    def _export(self) -> None:
        self.export_requested.emit()


def _card_downloads(card: ModCard) -> int:
    try:
        return int((card._dl.text() or "0").replace(",", ""))
    except Exception:  # noqa: BLE001
        return 0


def _card_today(card: ModCard) -> int:
    try:
        text = card._dl_cap.text()
        plus = text.split("+")[-1].split(" TODAY")[0].replace(",", "")
        return int(plus or 0)
    except Exception:  # noqa: BLE001
        return 0


def _card_growth(card: ModCard) -> int:
    text = card._growth.text()
    try:
        if "▲" in text:
            return int(text.split("+")[-1].replace(",", "").split()[0])
    except Exception:  # noqa: BLE001
        pass
    return -1 if "— no change" in text else 0


class HistoryPage(QWidget):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        title = QLabel("Download history")
        title.setObjectName("PageTitle")
        layout.addWidget(title)
        hint = QLabel("Snapshots over time for the selected item. Delta = change since previous snapshot.")
        hint.setObjectName("Hint")
        layout.addWidget(hint)

        row = QHBoxLayout()
        row.addWidget(QLabel("Item:"))
        self.mod_combo = QComboBox()
        self.mod_combo.currentIndexChanged.connect(lambda _: self._populate())
        row.addWidget(self.mod_combo, 1)
        self.reload_btn = QPushButton("Reload")
        self.reload_btn.clicked.connect(lambda: self.refresh(self._storage))
        row.addWidget(self.reload_btn)
        layout.addLayout(row)

        self.table = make_table(
            ["Fetched", "Downloads", "Today", "Delta", "Visits", "Visits today", "Rank", "Watchers"]
        )
        layout.addWidget(self.table, 1)
        self._storage = None

    def refresh(self, storage: Storage) -> None:
        self._storage = storage
        current = self.mod_combo.currentData()
        self.mod_combo.blockSignals(True)
        self.mod_combo.clear()
        for m in storage.get_mods(active_only=True):
            self.mod_combo.addItem(f"{m['name']}  ({m['content_type']})", int(m["id"]))
        if current is not None:
            idx = self.mod_combo.findData(current)
            if idx >= 0:
                self.mod_combo.setCurrentIndex(idx)
        self.mod_combo.blockSignals(False)
        self._populate()

    def _populate(self) -> None:
        if self._storage is None:
            return
        mod_id = self.mod_combo.currentData()
        if mod_id is None:
            fill_table(self.table, [])
            return
        rows: List[List[Any]] = []
        prev_total = None
        for s in self._storage.snapshots_for(mod_id):
            total = s["downloads_total"]
            delta = ""
            if prev_total is not None and total is not None:
                delta = max(0, int(total) - int(prev_total))
            prev_total = total
            rank = f"{s['rank'] or '-'}/{s['rank_total'] or '-'}" if s["rank"] is not None else "-"
            rows.append([
                s["fetched_at"],
                s["downloads_total"],
                s["downloads_today"],
                delta,
                s["visits"],
                s["visits_today"],
                rank,
                s["watchers"] or 0,
            ])
        fill_table(self.table, rows)


class CommentsPage(QWidget):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._filter = ""
        self._rows: List[List[Any]] = []
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        title = QLabel("Comments & replies")
        title.setObjectName("PageTitle")
        layout.addWidget(title)
        hint = QLabel("All comments seen on your tracked items (newest first).")
        hint.setObjectName("Hint")
        layout.addWidget(hint)

        toolbar = QHBoxLayout()
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["Filter: All", "Filter: Replies only", "Filter: New comments"])
        self.filter_combo.currentIndexChanged.connect(lambda _: self._apply_filter())
        self.search = QLineEdit()
        self.search.setObjectName("Search")
        self.search.setPlaceholderText("Search comments\u2026")
        self.search.setClearButtonEnabled(True)
        self.search.setFixedWidth(240)
        self.search.textChanged.connect(lambda _: self._apply_filter())
        self.open_btn = QPushButton("Open selected on ModDB")
        self.open_btn.clicked.connect(self._open_selected)
        toolbar.addWidget(self.filter_combo)
        toolbar.addWidget(self.search)
        toolbar.addStretch(1)
        toolbar.addWidget(self.open_btn)
        layout.addLayout(toolbar)

        self.table = make_table(["Posted", "Mod", "Author", "Content"])
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.doubleClicked.connect(lambda _: self._open_selected())
        layout.addWidget(self.table, 1)
        self.placeholder = QLabel("No comments yet. Poll your tracked items to gather them.")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setObjectName("Hint")
        self.placeholder.setMinimumHeight(120)
        layout.addWidget(self.placeholder)
        self._storage = None
        self._mod_urls: Dict[int, str] = {}

    def refresh(self, storage: Storage) -> None:
        self._storage = storage
        self._mod_urls = {}
        for m in storage.get_mods(active_only=False):
            m = dict(m)
            self._mod_urls[int(m["id"])] = str(m.get("url") or "")
        rows: List[List[Any]] = []
        for c in storage.recent_comments(300):
            rows.append([
                c["posted_at"],
                self._mod_name(storage, c["mod_id"]),
                c["author"],
                (c["content"] or "").replace("\n", " ")[:300],
                c["mod_id"],
                c["parent_id"],
            ])
        self._rows = rows
        self._apply_filter()

    @staticmethod
    def _mod_name(storage: Storage, mod_id: int) -> str:
        for m in storage.get_mods(active_only=False):
            if int(m["id"]) == int(mod_id):
                return m["name"]
        return str(mod_id)

    def apply_filter(self, text: str) -> None:
        self.search.setText(text)
        self._apply_filter()

    def _apply_filter(self) -> None:
        needle = self.search.text().strip().lower()
        mode = self.filter_combo.currentIndex()
        out = []
        for row in self._rows:
            is_reply = bool(row[5])
            if mode == 1 and is_reply:
                out.append(row)
            elif mode == 2 and not is_reply:
                out.append(row)
            elif mode == 0:
                out.append(row)
        if needle:
            out = [r for r in out if needle in f"{r[1]} {r[2]} {r[3]}".lower()]
        fill_table(self.table, [r[:4] for r in out])
        self.placeholder.setVisible(not self._rows)
        self.table.setVisible(bool(self._rows))

    def _open_selected(self) -> None:
        idx = self.table.currentRow()
        if idx < 0 or idx >= len(self._rows):
            return
        mod_id = self._rows[idx][4]
        url = self._mod_urls.get(int(mod_id), "")
        if url:
            QDesktopServices.openUrl(QUrl(url))


class EventsPage(QWidget):
    """Notification center: unseen highlighting, kind filter, mark-as-read, open on ModDB."""

    open_url = pyqtSignal(str)
    events_read = pyqtSignal()

    KIND_LABELS = {"download": "Downloads", "comment": "Comments", "reply": "Replies"}
    KIND_STYLE = {"download": ("⬇", SUCCESS), "comment": ("💬", ACCENT), "reply": ("↩", WARNING)}

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._filter = ""
        self._kind = ""
        self._rows: List[Dict[str, Any]] = []
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)
        title = QLabel("Notifications")
        title.setObjectName("PageTitle")
        layout.addWidget(title)
        hint = QLabel("Download and comment events detected by polls. Double-click a row to open it on ModDB.")
        hint.setObjectName("Hint")
        layout.addWidget(hint)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)
        self.search = QLineEdit()
        self.search.setObjectName("Search")
        self.search.setPlaceholderText("Search events\u2026")
        self.search.setClearButtonEnabled(True)
        self.search.setFixedWidth(240)
        self.search.textChanged.connect(lambda _: self._apply_filter())
        toolbar.addWidget(self.search)
        self.kind_combo = QComboBox()
        self.kind_combo.addItem("All kinds", "")
        self.kind_combo.addItem("Downloads", "download")
        self.kind_combo.addItem("Comments", "comment")
        self.kind_combo.addItem("Replies", "reply")
        self.kind_combo.currentIndexChanged.connect(lambda _: self._apply_filter())
        toolbar.addWidget(self.kind_combo)
        toolbar.addStretch(1)
        self.mark_read_btn = QPushButton("Mark all read")
        self.mark_read_btn.clicked.connect(self.mark_all_seen)
        toolbar.addWidget(self.mark_read_btn)
        layout.addLayout(toolbar)

        self.table = make_table(["Time", "Kind", "Mod", "Message"])
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._context_menu)
        self.table.itemDoubleClicked.connect(lambda _: self._open_selected())
        layout.addWidget(self.table, 1)
        self.placeholder = QLabel("No events yet. Run a poll to start recording activity.")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setObjectName("Hint")
        self.placeholder.setMinimumHeight(120)
        layout.addWidget(self.placeholder)

    def refresh(self, storage: Storage) -> None:
        self._storage = storage
        rows: List[Dict[str, Any]] = []
        for e in storage.recent_events(200):
            rows.append({
                "id": e["id"],
                "created_at": e["created_at"],
                "kind": e["kind"],
                "mod_name": e["mod_name"] or "-",
                "message": e["message"],
                "url": e["url"] or "",
                "seen": bool(e["seen"]),
            })
        self._rows = rows
        self._apply_filter()

    def mark_all_seen(self) -> None:
        storage = getattr(self, "_storage", None)
        if storage is not None:
            storage.mark_events_seen()
        for r in self._rows:
            r["seen"] = True
        self._apply_filter()
        self.events_read.emit()

    def apply_filter(self, text: str) -> None:
        self.search.setText(text)
        self._apply_filter()

    def _mark_all_read(self) -> None:
        self.mark_all_seen()

    def _open_selected(self) -> None:
        row = self.table.currentRow()
        if 0 <= row < len(self._displayed):
            url = self._displayed[row].get("url")
            if url:
                self.open_url.emit(url)

    def _context_menu(self, pos) -> None:
        row = self.table.rowAt(pos.y())
        if row < 0 or row >= len(self._displayed):
            return
        entry = self._displayed[row]
        menu = QMenu(self)
        if entry.get("url"):
            open_act = QAction("Open on ModDB", self)
            open_act.triggered.connect(lambda: self.open_url.emit(entry["url"]))
            menu.addAction(open_act)
            copy_act = QAction("Copy URL", self)
            copy_act.triggered.connect(lambda: QApplication.clipboard().setText(entry["url"]))
            menu.addAction(copy_act)
            menu.addSeparator()
        mark_act = QAction("Mark all read", self)
        mark_act.triggered.connect(self._mark_all_read)
        menu.addAction(mark_act)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _apply_filter(self) -> None:
        needle = self.search.text().strip().lower()
        self._kind = self.kind_combo.currentData() or ""
        out = []
        for r in self._rows:
            if self._kind and r["kind"] != self._kind:
                continue
            if needle and needle not in f"{r['kind']} {r['mod_name']} {r['message']}".lower():
                continue
            out.append(r)
        self._displayed = out

        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(out))
        for r, entry in enumerate(out):
            icon, color = self.KIND_STYLE.get(entry["kind"], ("•", GRAY))
            kind_label = self.KIND_LABELS.get(entry["kind"], entry["kind"].title())
            cells = [entry["created_at"], f"{icon} {kind_label}", entry["mod_name"], entry["message"]]
            for c, value in enumerate(cells):
                item = QTableWidgetItem(str(value))
                if not entry["seen"]:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    item.setForeground(QColor(TEXT))
                if c == 0:
                    item.setData(Qt.ItemDataRole.UserRole, entry["created_at"])
                self.table.setItem(r, c, item)
        self.table.resizeColumnsToContents()

        self.placeholder.setVisible(not self._rows)
        self.table.setVisible(bool(self._rows))


class LogPage(QPlainTextEdit):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setReadOnly(True)
        self.setMaximumBlockCount(3000)
        self.setStyleSheet(
            f"QPlainTextEdit {{ background: {PANEL}; color: {GRAY}; border: 1px solid {BORDER};"
            "border-radius: 8px; font-family: Consolas, monospace; font-size: 12px; }}"
        )


def _merge_defaults(data: Dict[str, Any]) -> Dict[str, Any]:
    """Deep-merge a parsed JSON object over the built-in default config."""
    base = json.loads(json.dumps(tracker.DEFAULT_CONFIG))

    def _merge(b: Dict[str, Any], o: Dict[str, Any]) -> None:
        for key, value in o.items():
            if isinstance(value, dict) and isinstance(b.get(key), dict):
                _merge(b[key], value)
            else:
                b[key] = value

    _merge(base, data)
    return base


class SettingsPage(QWidget):
    saved = pyqtSignal()
    file_reload = pyqtSignal()

    def __init__(self, config: Dict[str, Any], parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._config = config

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        title = QLabel("Configuration")
        title.setObjectName("PageTitle")
        root.addWidget(title)
        hint = QLabel("Everything is edited in-app and saved to config.json.")
        hint.setObjectName("Hint")
        root.addWidget(hint)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setSpacing(10)

        # ---- profile
        layout.addWidget(section_label("Member profile"))
        box = panel(content)
        gl = QGridLayout(box)
        gl.setHorizontalSpacing(12)
        gl.setVerticalSpacing(10)
        gl.addWidget(QLabel("Profile URL:"), 0, 0)
        self.profile = QLineEdit()
        self.profile.setPlaceholderText("https://www.moddb.com/members/yourname")
        gl.addWidget(self.profile, 0, 1)
        self.auto_discover = QCheckBox("Auto-discover mods/addons/files from the profile")
        gl.addWidget(self.auto_discover, 1, 0, 1, 2)
        layout.addWidget(box)

        # ---- manual mods
        layout.addWidget(section_label("Manual mods (tracked in addition to auto-discovered)"))
        box2 = panel(content)
        ml = QVBoxLayout(box2)
        self.mods_list = QListWidget()
        ml.addWidget(self.mods_list)
        mrow = QHBoxLayout()
        self.mod_url_input = QLineEdit()
        self.mod_url_input.setPlaceholderText("https://www.moddb.com/mods/...")
        add_btn = QPushButton("Add")
        add_btn.clicked.connect(self._add_mod)
        remove_btn = QPushButton("Remove selected")
        remove_btn.clicked.connect(self._remove_mod)
        mrow.addWidget(self.mod_url_input, 1)
        mrow.addWidget(add_btn)
        mrow.addWidget(remove_btn)
        ml.addLayout(mrow)
        layout.addWidget(box2)

        # ---- polling
        layout.addWidget(section_label("Polling"))
        box3 = panel(content)
        pl = QGridLayout(box3)
        pl.setHorizontalSpacing(12)
        pl.addWidget(QLabel("Interval (minutes):"), 0, 0)
        self.interval = QSpinBox()
        self.interval.setRange(5, 1440)
        self.interval.setSuffix(" min")
        pl.addWidget(self.interval, 0, 1)
        self.charts_each_poll = QCheckBox("Regenerate charts after each poll")
        pl.addWidget(self.charts_each_poll, 1, 0, 1, 2)
        layout.addWidget(box3)

        # ---- notifications
        layout.addWidget(section_label("Notifications"))
        box4 = panel(content)
        nl = QVBoxLayout(box4)
        self.notify_downloads = QCheckBox("New downloads")
        self.notify_comments = QCheckBox("New comments")
        self.notify_replies = QCheckBox("Replies to me")
        nl.addWidget(self.notify_downloads)
        nl.addWidget(self.notify_comments)
        nl.addWidget(self.notify_replies)
        ng = QGridLayout()
        ng.addWidget(QLabel("Max toasts per poll:"), 0, 0)
        self.max_toasts = QSpinBox()
        self.max_toasts.setRange(1, 20)
        ng.addWidget(self.max_toasts, 0, 1)
        ng.addWidget(QLabel("Toast app id:"), 1, 0)
        self.app_id = QLineEdit()
        ng.addWidget(self.app_id, 1, 1)
        nl.addLayout(ng)
        layout.addWidget(box4)

        # ---- tray
        layout.addWidget(section_label("System tray"))
        box5 = panel(content)
        tl = QVBoxLayout(box5)
        self.close_to_tray = QCheckBox("Minimize to tray when closing the window")
        self.start_minimized = QCheckBox("Start hidden in the tray (background mode)")
        tl.addWidget(self.close_to_tray)
        tl.addWidget(self.start_minimized)
        th = QLabel("The app keeps polling and showing toasts while it sits in the tray.")
        th.setObjectName("Hint")
        tl.addWidget(th)
        layout.addWidget(box5)

        # ---- paths
        layout.addWidget(section_label("Paths"))
        box6 = panel(content)
        pg = QGridLayout(box6)
        pg.setHorizontalSpacing(12)
        pg.addWidget(QLabel("Database:"), 0, 0)
        self.path_db = QLineEdit()
        pg.addWidget(self.path_db, 0, 1)
        pg.addWidget(QLabel("Charts output:"), 1, 0)
        self.path_output = QLineEdit()
        pg.addWidget(self.path_output, 1, 1)
        pg.addWidget(QLabel("Logs:"), 2, 0)
        self.path_logs = QLineEdit()
        pg.addWidget(self.path_logs, 2, 1)
        layout.addWidget(box6)

        # ---- advanced JSON
        layout.addWidget(section_label("Advanced"))
        box7 = panel(content)
        al = QVBoxLayout(box7)
        self.json_toggle = QCheckBox("Edit config.json as raw JSON")
        self.json_toggle.toggled.connect(self._toggle_json)
        al.addWidget(self.json_toggle)
        self.json_editor = QPlainTextEdit()
        self.json_editor.setMaximumHeight(240)
        self.json_editor.setStyleSheet(
            f"QPlainTextEdit {{ background: {PANEL2}; color: {TEXT}; border: 1px solid {BORDER};"
            "border-radius: 6px; font-family: Consolas, monospace; font-size: 12px; }}"
        )
        al.addWidget(self.json_editor)
        self.json_editor.hide()
        aj = QHBoxLayout()
        self.apply_json_btn = QPushButton("Apply JSON to form")
        self.apply_json_btn.clicked.connect(self._apply_json)
        aj.addWidget(self.apply_json_btn)
        aj.addStretch(1)
        al.addLayout(aj)
        self.apply_json_btn.hide()
        layout.addWidget(box7)

        layout.addStretch(1)

        # ---- buttons
        buttons = QHBoxLayout()
        self.save_btn = QPushButton("Save configuration")
        self.save_btn.setObjectName("Primary")
        self.save_btn.clicked.connect(self._save)
        self.reload_btn = QPushButton("Reload from file")
        self.reload_btn.clicked.connect(self._reload_from_file)
        buttons.addWidget(self.save_btn)
        buttons.addWidget(self.reload_btn)
        buttons.addStretch(1)
        layout.addLayout(buttons)

        scroll.setWidget(content)
        root.addWidget(scroll, 1)

        self.reload()

    # ---- helpers --------------------------------------------------------
    def _add_mod(self) -> None:
        url = self.mod_url_input.text().strip()
        if not url:
            return
        existing = [self.mods_list.item(i).text() for i in range(self.mods_list.count())]
        if url not in existing:
            self.mods_list.addItem(url)
        self.mod_url_input.clear()

    def _remove_mod(self) -> None:
        for item in self.mods_list.selectedItems():
            self.mods_list.takeItem(self.mods_list.row(item))

    def _toggle_json(self, checked: bool) -> None:
        self.json_editor.setVisible(checked)
        self.apply_json_btn.setVisible(checked)
        if checked:
            self.json_editor.setPlainText(json.dumps(self.current_config(), indent=2))

    def _apply_json(self) -> None:
        try:
            parsed = json.loads(self.json_editor.toPlainText())
            if not isinstance(parsed, dict):
                raise ValueError("Root value must be a JSON object")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Invalid JSON", str(exc))
            return
        merged = _merge_defaults(parsed)
        self._config = merged
        self.reload()
        self.json_editor.setPlainText(json.dumps(merged, indent=2))
        QMessageBox.information(self, "Applied", "JSON applied to the form. Review and press Save.")

    def _reload_from_file(self) -> None:
        try:
            self._config = load_config(CONFIG_FILE)
            self.reload()
            self.file_reload.emit()
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Reload failed", str(exc))

    # ---- data -----------------------------------------------------------
    def reload(self) -> None:
        self.profile.setText(self._config.get("profile_url", ""))
        self.auto_discover.setChecked(bool(self._config.get("auto_discover", True)))
        self.mods_list.clear()
        for url in self._config.get("mods", []):
            if url:
                self.mods_list.addItem(url)
        self.interval.setValue(int(self._config["poll"].get("interval_minutes", 30)))
        self.charts_each_poll.setChecked(bool(self._config["poll"].get("charts_each_poll", True)))
        self.notify_downloads.setChecked(bool(self._config["poll"].get("notify_on_downloads", True)))
        self.notify_comments.setChecked(bool(self._config["poll"].get("notify_on_comments", True)))
        self.notify_replies.setChecked(bool(self._config["poll"].get("notify_on_replies", True)))
        self.max_toasts.setValue(int(self._config["notifications"].get("max_toasts", 5)))
        self.app_id.setText(self._config["notifications"].get("app_id", "ModDB Tracker"))
        self.close_to_tray.setChecked(bool(self._config.get("tray", {}).get("close_to_tray", True)))
        self.start_minimized.setChecked(bool(self._config.get("tray", {}).get("start_minimized", False)))
        self.path_db.setText(self._config["paths"].get("db", "tracker.db"))
        self.path_output.setText(self._config["paths"].get("output", "output"))
        self.path_logs.setText(self._config["paths"].get("logs", "logs"))
        if self.json_editor.isVisible():
            self.json_editor.setPlainText(json.dumps(self.current_config(), indent=2))

    def current_config(self) -> Dict[str, Any]:
        cfg = json.loads(json.dumps(self._config))
        cfg["profile_url"] = self.profile.text().strip()
        cfg["auto_discover"] = self.auto_discover.isChecked()
        cfg["mods"] = [self.mods_list.item(i).text().strip() for i in range(self.mods_list.count()) if self.mods_list.item(i).text().strip()]
        cfg["poll"]["interval_minutes"] = self.interval.value()
        cfg["poll"]["charts_each_poll"] = self.charts_each_poll.isChecked()
        cfg["poll"]["notify_on_downloads"] = self.notify_downloads.isChecked()
        cfg["poll"]["notify_on_comments"] = self.notify_comments.isChecked()
        cfg["poll"]["notify_on_replies"] = self.notify_replies.isChecked()
        cfg["notifications"]["max_toasts"] = self.max_toasts.value()
        cfg["notifications"]["app_id"] = self.app_id.text().strip() or "ModDB Tracker"
        cfg["tray"]["close_to_tray"] = self.close_to_tray.isChecked()
        cfg["tray"]["start_minimized"] = self.start_minimized.isChecked()
        cfg["paths"]["db"] = self.path_db.text().strip() or "tracker.db"
        cfg["paths"]["output"] = self.path_output.text().strip() or "output"
        cfg["paths"]["logs"] = self.path_logs.text().strip() or "logs"
        return cfg

    def _save(self) -> None:
        try:
            save_config(self.current_config(), CONFIG_FILE)
            QMessageBox.information(self, "Saved", "Configuration saved to config.json")
            self.saved.emit()
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Save failed", str(exc))


# --------------------------------------------------------------------------
# main window
# --------------------------------------------------------------------------

class TrackerWindow(QMainWindow):
    NAV = [
        ("Dashboard", "🏠"),
        ("My Mods", "📦"),
        ("Analytics", "📊"),
        ("Compare", "⚖️"),
        ("Insights", "💡"),
        ("Achievements", "🏆"),
        ("History", "📈"),
        ("Comments", "💬"),
        ("Notifications", "🔔"),
        ("Configuration", "⚙️"),
        ("Log", "🪵"),
    ]

    def __init__(self, config_path: str = CONFIG_FILE) -> None:
        super().__init__()
        self.config_path = config_path
        self.config = load_config(config_path)
        transport.patch_moddb()  # ensure moddb's internal get_page is Cloudflare-proof
        self.storage = Storage(self.config["paths"]["db"])
        self.worker: Optional[TrackerWorker] = None
        self._stopping = False
        self._quit_requested = False
        self._tray_hint_shown = False
        self.tray: Optional[QSystemTrayIcon] = None

        self.setWindowTitle("ModDB Tracker")
        self.resize(1280, 820)
        self.setMinimumSize(1080, 680)

        self._search_override = False
        self._saved_page = 0
        self._build_ui()
        self._wire_logging()
        self._build_tray()

        self.auto_timer = QTimer(self)
        self.auto_timer.timeout.connect(self._poll_now)
        self.auto_checkbox = QCheckBox("Auto-poll")
        self.auto_checkbox.toggled.connect(self._toggle_auto_poll)
        self.statusBar().addWidget(self.auto_checkbox)
        self.busy_bar = QProgressBar()
        self.busy_bar.setRange(0, 0)
        self.busy_bar.setFixedWidth(120)
        self.busy_bar.setVisible(False)
        self.statusBar().addWidget(self.busy_bar)

        self.status_label = QLabel("")
        self.statusBar().addPermanentWidget(self.status_label)

        self._db_label = QLabel("")
        self.statusBar().addPermanentWidget(self._db_label)

        self.refresh_all()
        self.set_status("Ready. Click \u201cPoll now\u201d to fetch data.")

    # ---- ui construction ------------------------------------------------
    def _build_ui(self) -> None:
        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_top_bar())

        body = QHBoxLayout()
        body.setContentsMargins(14, 14, 14, 12)
        body.setSpacing(14)
        body.addWidget(self._build_sidebar())

        self.stack = QStackedWidget()
        self.dashboard = DashboardPage(self.config, self.config_path)
        self.mods = ModsPage(self.config, self.config_path)
        self.analytics = AnalyticsPage(self.config, self.config_path)
        self.compare = ComparePage(self.config, self.config_path)
        self.insights = InsightsPage()
        self.achievements = AchievementsPage()
        self.history = HistoryPage()
        self.comments = CommentsPage()
        self.events = EventsPage()
        self.log_page = LogPage()
        self.settings = SettingsPage(self.config)
        self.search_page = SearchResultsPage()
        for page in (self.dashboard, self.mods, self.analytics, self.compare, self.insights, self.achievements, self.history, self.comments, self.events, self.settings, self.log_page):
            self.stack.addWidget(page)
        self.stack.addWidget(self.search_page)
        self.search_page.open_url.connect(self._open_url)
        self.settings.saved.connect(self._on_settings_saved)
        self.settings.file_reload.connect(self._on_settings_saved)
        self.mods.open_requested.connect(self._open_url)
        self.mods.refresh_requested.connect(self._refresh_mod)
        self.mods.favorite_toggled.connect(self._toggle_favorite)
        self.mods.remove_requested.connect(self._remove_mod)
        self.mods.export_requested.connect(self._export_json)
        self.dashboard.regen_requested.connect(self._regenerate_charts)
        self.dashboard.view_insights.connect(self._open_insights)
        self.events.open_url.connect(self._open_url)
        self.events.events_read.connect(self._update_badge)
        body.addWidget(self.stack, 1)
        root.addLayout(body, 1)

        self.setCentralWidget(central)
        self.sidebar.setCurrentRow(0)

    def _build_top_bar(self) -> QWidget:
        bar = QFrame()
        bar.setObjectName("TopBar")
        h = QHBoxLayout(bar)
        h.setContentsMargins(18, 10, 18, 10)
        h.setSpacing(12)

        logo = QLabel()
        logo.setPixmap(QIcon(tray_icon_path()).pixmap(30, 30))
        logo.setFixedSize(30, 30)
        title = QLabel("ModDB Tracker")
        title.setStyleSheet("font-size: 16px; font-weight: 800; letter-spacing: 0.5px;")
        version = QLabel(f"v{VERSION}")
        version.setStyleSheet(
            f"color: {ACCENT}; font-size: 10px; font-weight: 700; background: {ACCENT}22;"
            "padding: 2px 7px; border-radius: 8px;"
        )
        brand = QHBoxLayout()
        brand.setSpacing(8)
        brand.addWidget(logo)
        brand.addWidget(title)
        brand.addWidget(version, 0, Qt.AlignmentFlag.AlignVCenter)
        h.addLayout(brand)

        h.addStretch(1)

        self.search_box = QLineEdit()
        self.search_box.setObjectName("Search")
        self.search_box.setPlaceholderText("Search mods, comments, events\u2026")
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setFixedWidth(240)
        self.search_box.textChanged.connect(self._search_changed)
        h.addWidget(self.search_box)

        self.sync_label = QLabel("Not synced")
        self.sync_label.setStyleSheet(f"color: {GRAY}; font-size: 12px;")
        h.addWidget(self.sync_label)

        self.bell_btn = QToolButton()
        self.bell_btn.setText("🔔")
        self.bell_btn.setToolTip("Notifications")
        self.bell_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.bell_btn.clicked.connect(self._show_notifications)
        h.addWidget(self.bell_btn)
        self.badge = QLabel("")
        self.badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.badge.setMinimumWidth(18)
        self.badge.setStyleSheet(
            f"background: {ERROR}; color: #FFFFFF; font-size: 10px; font-weight: 700;"
            "border-radius: 9px; padding: 0 5px;"
        )
        self.badge.setVisible(False)
        h.addWidget(self.badge)

        self.export_btn = QToolButton()
        self.export_btn.setText("Export \u25be")
        self.export_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.export_btn.setToolTip("Export reports")
        export_menu = QMenu(self)
        export_menu.addAction("CSV\u2026", self._export_csv)
        export_menu.addAction("Excel (.xlsx)\u2026", self._export_xlsx)
        export_menu.addAction("PDF report\u2026", self._export_pdf)
        export_menu.addSeparator()
        export_menu.addAction("JSON\u2026", self._export_json)
        self.export_btn.setMenu(export_menu)
        h.addWidget(self.export_btn)

        self.poll_btn = QPushButton("Poll now")
        self.poll_btn.setObjectName("Primary")
        self.poll_btn.clicked.connect(self._poll_now)
        self.rescan_btn = QPushButton("Rescan")
        self.rescan_btn.clicked.connect(self._rescan)
        self.charts_btn = QPushButton("Charts")
        self.charts_btn.clicked.connect(self._regenerate_charts)
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.clicked.connect(self.refresh_all)
        h.addWidget(self.poll_btn)
        h.addWidget(self.rescan_btn)
        h.addWidget(self.charts_btn)
        h.addWidget(self.refresh_btn)
        return bar

    def _build_sidebar(self) -> QWidget:
        self.sidebar = QListWidget()
        self.sidebar.setObjectName("Sidebar")
        self.sidebar.setFixedWidth(200)
        for name, icon in self.NAV:
            item = QListWidgetItem(f"{icon}   {name}")
            item.setSizeHint(QSize(0, 38))
            self.sidebar.addItem(item)
        self.sidebar.currentRowChanged.connect(self._nav_changed)
        return self.sidebar

    def _nav_changed(self, row: int) -> None:
        if self._search_override:
            self._search_override = False
            self.search_box.blockSignals(True)
            self.search_box.clear()
            self.search_box.blockSignals(False)
        self.stack.setCurrentIndex(row)
        if self.stack.currentWidget() is self.events:
            self.events.mark_all_seen()

    # ---- system tray ----------------------------------------------------
    def _build_tray(self) -> None:
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return
        self.tray = QSystemTrayIcon(QIcon(tray_icon_path()), self)
        self.tray.setToolTip("ModDB Tracker")

        menu = QMenu()
        self.tray_menu_show = QAction("Show / Hide", self)
        self.tray_menu_show.triggered.connect(self._toggle_window)
        self.tray_menu_poll = QAction("Poll now", self)
        self.tray_menu_poll.triggered.connect(self._poll_now)
        self.tray_menu_rescan = QAction("Rescan profile", self)
        self.tray_menu_rescan.triggered.connect(self._rescan)
        self.tray_menu_quit = QAction("Quit", self)
        self.tray_menu_quit.triggered.connect(self._request_quit)
        menu.addAction(self.tray_menu_show)
        menu.addAction(self.tray_menu_poll)
        menu.addAction(self.tray_menu_rescan)
        menu.addSeparator()
        menu.addAction(self.tray_menu_quit)
        self.tray.setContextMenu(menu)
        self.tray.activated.connect(self._tray_activated)
        self.tray.show()

    def _tray_activated(self, reason) -> None:
        if reason == QSystemTrayIcon.ActivationReason.Trigger:
            self._toggle_window()

    def _toggle_window(self) -> None:
        if self.isVisible():
            self.hide()
        else:
            self.show()
            self.showNormal()
            self.raise_()
            self.activateWindow()

    def hide_to_tray(self, first_time: bool = False) -> None:
        if self.tray is None:
            return
        self.hide()
        if first_time or not self._tray_hint_shown:
            self._tray_hint_shown = True
            self.tray.showMessage(
                "ModDB Tracker",
                "Still running in the background. Polling and notifications stay active.",
                QSystemTrayIcon.MessageIcon.Information,
                3000,
            )

    def _request_quit(self) -> None:
        self._quit_requested = True
        if self.tray is not None:
            self.tray.hide()
        self.close()

    def closeEvent(self, event) -> None:  # noqa: N802
        tray_active = self.tray is not None and self.config.get("tray", {}).get("close_to_tray", True)
        if not self._quit_requested and tray_active:
            event.ignore()
            self.hide_to_tray()
            return
        self._stopping = True
        self.auto_timer.stop()
        if self.worker is not None and self.worker.isRunning():
            self.worker.wait(5000)
        self.storage.close()
        event.accept()

    # ---- logging --------------------------------------------------------
    def _wire_logging(self) -> None:
        self._bridge = LogBridge(self)
        self._bridge.line.connect(self.log_page.appendPlainText)
        handler = GuiLogHandler(self._bridge)
        handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s"))
        logging.getLogger().addHandler(handler)

    # ---- data -----------------------------------------------------------
    def refresh_all(self) -> None:
        self.dashboard.refresh(self.storage)
        self.mods.refresh(self.storage)
        self.analytics.refresh(self.storage)
        self.compare.refresh(self.storage)
        self.insights.refresh(self.storage)
        self.achievements.refresh(self.storage)
        self.history.refresh(self.storage)
        self.comments.refresh(self.storage)
        self.events.refresh(self.storage)
        self.settings.reload()
        self._update_badge()

        member = self.storage.meta_get("member_name") or ""
        last = self.storage.meta_get("last_poll") or ""
        bits = [f"Member: {member}"] if member else ["Member: not set"]
        bits.append(f"Last poll: {last}" if last else "Never polled")
        self.status_label.setText("   |   ".join(bits))
        self.sync_label.setText(f"Synced {relative_time(last)}" if last else "Not synced")
        self._update_tray_tooltip(member, last)

        try:
            db_path = Path(self.storage.db_path) if hasattr(self.storage, "db_path") else Path(self.config["paths"]["db"])
            size_mb = db_path.stat().st_size / (1024 * 1024) if db_path.exists() else 0
            self._db_label.setText(f"DB {size_mb:.1f} MB")
        except Exception:  # noqa: BLE001
            pass

    def _on_settings_saved(self) -> None:
        self.config = load_config(self.config_path)
        self.storage = Storage(self.config["paths"]["db"])
        self.dashboard._config = self.config
        self.dashboard._load_prefs()
        self.dashboard._apply_layout()
        self.mods._config = self.config
        self.analytics._config = self.config
        self.compare._config = self.config
        self._restart_auto_timer()
        self.refresh_all()

    def _open_url(self, url: str) -> None:
        if url:
            QDesktopServices.openUrl(QUrl(url))

    # ---- actions --------------------------------------------------------
    def _set_busy(self, busy: bool, label: str = "") -> None:
        self.poll_btn.setEnabled(not busy)
        self.rescan_btn.setEnabled(not busy)
        self.charts_btn.setEnabled(not busy)
        self.refresh_btn.setEnabled(not busy)
        self.export_btn.setEnabled(not busy)
        self.dashboard.regen_btn.setEnabled(not busy)
        self.busy_bar.setVisible(busy)
        if busy:
            self.set_status(label or "Working...")

    def _update_badge(self) -> None:
        count = self.storage.count_unseen()
        self.badge.setText(f"{count}" if count < 100 else "99+")
        self.badge.setVisible(count > 0)
        self.bell_btn.setToolTip(f"{count} new notification(s)" if count else "Notifications")

    def _show_notifications(self) -> None:
        row = next(i for i, (name, _) in enumerate(self.NAV) if name == "Notifications")
        self.sidebar.setCurrentRow(row)
        self.stack.setCurrentIndex(row)
        self.events.mark_all_seen()

    def _open_insights(self) -> None:
        row = next(i for i, (name, _) in enumerate(self.NAV) if name == "Insights")
        self.sidebar.setCurrentRow(row)
        self.stack.setCurrentIndex(row)

    def set_status(self, text: str) -> None:
        self.status_label.setText(text)
        if self.tray is not None:
            self.tray.setToolTip(f"ModDB Tracker\n{text}")

    def _update_tray_tooltip(self, member: str, last: str) -> None:
        if self.tray is None:
            return
        line = f"Member: {member or 'not set'}   Last poll: {last or 'never'}"
        self.tray.setToolTip(f"ModDB Tracker\n{line}")

    def _on_worker_done(self, message: str) -> None:
        self._set_busy(False)
        self.refresh_all()
        self.set_status("Done.")
        if message:
            logger.info("Worker finished: %s", message)

    def _on_worker_failed(self, message: str) -> None:
        self._set_busy(False)
        self.refresh_all()
        self.set_status("Failed.")
        QMessageBox.critical(self, "Task failed", message)

    def _start_worker(self, fn: Callable, label: str) -> None:
        if self.worker is not None and self.worker.isRunning():
            return
        self._set_busy(True, label)
        self.worker = TrackerWorker(self.config, fn, parent=self)
        self.worker.done.connect(self._on_worker_done)
        self.worker.failed.connect(self._on_worker_failed)
        self.worker.start()

    def _poll_now(self) -> None:
        self._start_worker(run_poll, "Polling ModDB...")

    def _rescan(self) -> None:
        self._start_worker(run_discover, "Scanning profile...")

    def _regenerate_charts(self) -> None:
        def _charts(storage: Storage, config: Dict[str, Any]) -> Any:
            import charts

            paths = charts.generate_all(storage, Path(config["paths"]["output"]))
            return f"{len(paths)} chart(s) written"

        self._start_worker(_charts, "Regenerating charts...")

    # ---- per-mod actions ------------------------------------------------
    def _refresh_mod(self, mod_id: int) -> None:
        mod = next((dict(m) for m in self.storage.get_mods(active_only=True) if int(m["id"]) == int(mod_id)), None)
        if mod is None:
            return

        def _one(storage: Storage, config: Dict[str, Any]) -> str:
            target = next((dict(m) for m in storage.get_mods(active_only=True) if int(m["id"]) == int(mod_id)), None)
            if target is None:
                return "mod not found"
            tracker.snapshot_mod(storage, target, config, notify=False)
            return f"Refreshed {target['name']}"

        self._start_worker(_one, f"Refreshing {mod['name']}...")

    def _toggle_favorite(self, name_id: str, favorite: bool) -> None:
        self.storage.set_mod_favorite(name_id, favorite)
        self.refresh_all()

    def _remove_mod(self, name_id: str) -> None:
        mod = next((dict(m) for m in self.storage.get_mods(active_only=False) if str(m["name_id"]) == name_id), None)
        name = mod["name"] if mod else name_id
        answer = QMessageBox.question(
            self,
            "Remove from tracking",
            f"Stop tracking \u201c{name}\u201d?\nExisting history is kept, it just won't be polled anymore.",
        )
        if answer == QMessageBox.StandardButton.Yes:
            self.storage.set_mod_active(name_id, False)
            self.refresh_all()

    def _pick_save_path(self, title: str, default: str, flt: str) -> Optional[str]:
        path, _ = QFileDialog.getSaveFileName(self, title, default, flt)
        return path or None

    def _readonly_storage(self) -> Storage:
        return Storage(self.config["paths"]["db"])

    def _export_json(self) -> None:
        path = self._pick_save_path("Export data", "moddb_tracker_export.json", "JSON (*.json)")
        if not path:
            return
        try:
            storage = self._readonly_storage()
            try:
                storage.export_json(path)
            finally:
                storage.close()
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    def _export_csv(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "Choose a folder for CSV export")
        if not directory:
            return
        try:
            storage = self._readonly_storage()
            try:
                meta = {int(m["id"]): dict(m) for m in storage.get_mods(active_only=True)}
                csv_rows = []
                for m in storage.totals_per_mod():
                    mid = int(m["id"])
                    csv_rows.append([m["name"], m["downloads_total"], m["downloads_today"], m["visits"],
                                     m["watchers"], m["rating"], m["rank"],
                                     meta.get(mid, {}).get("content_type", "mod")])
                self._write_csv_rows(Path(directory) / "mods.csv",
                                     ["Name", "Downloads", "Today", "Visits", "Watchers", "Rating", "Rank", "Content type"],
                                     csv_rows)
                self._write_csv_rows(Path(directory) / "comments.csv",
                                     ["Posted", "Mod", "Author", "Text"],
                                     [list(r) for r in self._comment_list(storage)])
                self._write_csv_rows(Path(directory) / "events.csv",
                                     ["Time", "Kind", "Mod", "Message"],
                                     [[e["created_at"], e["kind"], e["mod_name"], e["message"]] for e in storage.recent_events(500)])
                self._write_csv_rows(Path(directory) / "history.csv",
                                     ["Mod", "Time", "Downloads", "Visits", "Watchers"],
                                     [[m["name"], s["fetched_at"], s["downloads_total"],
                                       s["visits"], s["watchers"]]
                                      for m in storage.get_mods(active_only=True)
                                      for s in storage.snapshots_for(int(m["id"]))][-2000:])
            finally:
                storage.close()
            QMessageBox.information(self, "CSV export", "Saved to:\n" + str(directory))
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    @staticmethod
    def _comment_list(storage) -> List[Tuple[str, str, str, Any]]:
        mod_names = {int(m["id"]): m["name"] for m in storage.get_mods(active_only=False)}
        return [(c["posted_at"], mod_names.get(int(c["mod_id"]), "-"), c["author"], c["content"])
                for c in storage.recent_comments(500)]

    @staticmethod
    def _write_csv_rows(path: Path, headers: List[str], rows: List[List[Any]]) -> None:
        import csv

        with open(str(path), "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            writer.writerows(rows)

    def _export_xlsx(self) -> None:
        path = self._pick_save_path("Export to Excel", "moddb_tracker.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            storage = self._readonly_storage()
            try:
                wb = Workbook()
                wb.remove(wb.active)
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="3B82F6")
                meta = {int(m["id"]): dict(m) for m in storage.get_mods(active_only=True)}
                xlsx_mods = []
                for m in storage.totals_per_mod():
                    mid = int(m["id"])
                    xlsx_mods.append([m["name"], m["downloads_total"], m["downloads_today"], m["visits"], m["watchers"],
                                      m["rating"], m["rank"], meta.get(mid, {}).get("content_type", "mod")])
                sections = {
                    "Mods": (["Name", "Downloads", "Today", "Visits", "Watchers", "Rating", "Rank", "Content type"],
                             xlsx_mods),
                    "Comments": (["Posted", "Mod", "Author", "Text"],
                                 [list(r) for r in self._comment_list(storage)]),
                    "Events": (["Time", "Kind", "Mod", "Message"],
                               [[e["created_at"], e["kind"], e["mod_name"], e["message"]] for e in storage.recent_events(500)]),
                    "History": (["Mod", "Time", "Downloads", "Visits", "Watchers"],
                                [[m["name"], s["fetched_at"], s["downloads_total"],
                                  s["visits"], s["watchers"]]
                                 for m in storage.get_mods(active_only=True)
                                 for s in storage.snapshots_for(int(m["id"]))][-2000:]),
                }
                for name, (headers, rows) in sections.items():
                    ws = wb.create_sheet(name)
                    ws.append(headers)
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                    for row in rows:
                        ws.append(row)
                    for col, _ in enumerate(headers, start=1):
                        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(12, min(45, len(str(headers[col - 1])) + 2))
                wb.save(path)
            finally:
                storage.close()
            QMessageBox.information(self, "Excel export", f"Saved to:\n{path}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    def _export_pdf(self) -> None:
        path = self._pick_save_path("Export PDF report", "moddb_tracker_report.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            storage = self._readonly_storage()
            try:
                summary = analytics.all_mods_summary(storage, 30)
                member = storage.meta_get("member_name") or "—"
                last_poll = storage.meta_get("last_poll") or "—"

                styles = getSampleStyleSheet()
                title_style = ParagraphStyle("TitleX", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#0F172A"))
                body = ParagraphStyle("BodyX", parent=styles["BodyText"], fontSize=9, leading=12)
                small = ParagraphStyle("SmallX", parent=styles["BodyText"], fontSize=8, textColor=colors.HexColor("#64748B"))

                doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
                                        topMargin=18 * mm, bottomMargin=18 * mm, title="ModDB Tracker report")
                story = [Paragraph("ModDB Tracker — weekly report", title_style),
                         Spacer(1, 4),
                         Paragraph(f"Member: {member} &nbsp;&nbsp;|&nbsp;&nbsp; Last poll: {last_poll} &nbsp;&nbsp;|&nbsp;&nbsp; "
                                   f"Generated: {datetime.datetime.now():%Y-%m-%d %H:%M}", small),
                         Spacer(1, 10)]

                agg = analytics.aggregate_summary(summary)
                story.append(Paragraph(
                    f"<b>Aggregate</b> — {agg['count']} tracked mods, {agg['total']:,} total downloads, "
                    f"<b>+{agg['delta_7d']:,}</b> in the last 7 days, <b>+{agg['delta_30d']:,}</b> in 30, "
                    f"~<b>{agg['next_week']:,}</b> projected for next week.", body))
                story.append(Spacer(1, 8))

                rows = [["Mod", "Total", "7d", "30d", "Avg/day", "Best day", "Next week"]]
                for s in summary:
                    best = s["best_day"]["label"] if s["best_day"] else "—"
                    nxt = f"{s['next_week_estimate']:,}" if s["next_week_estimate"] else "—"
                    rows.append([Paragraph(s["name"], body), f"{s['total']:,}", f"+{s['delta_7d']:,}",
                                 f"+{s['delta_30d']:,}", f"{s['avg_per_day']}", best, nxt])
                table = Table(rows, colWidths=[70 * mm, 22 * mm, 18 * mm, 18 * mm, 20 * mm, 28 * mm, 22 * mm], repeatRows=1)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]))
                story.append(table)
                story.append(Spacer(1, 10))

                comments = self._comment_list(storage)
                if comments:
                    story.append(Paragraph("<b>Recent comments</b>", body))
                    story.append(Spacer(1, 4))
                    for posted, mod_name, author, content in comments[:20]:
                        text = " ".join(str(content).split())
                        story.append(Paragraph(f"<b>{author}</b> on {mod_name} ({posted}): {text[:220]}", body))
                    story.append(Spacer(1, 8))

                events = storage.recent_events(20)
                if events:
                    story.append(Paragraph("<b>Recent events</b>", body))
                    story.append(Spacer(1, 4))
                    for e in events:
                        story.append(Paragraph(f"{e['created_at']} — {e['kind']} · {e['mod_name']}: {e['message']}", small))
                doc.build(story)
            finally:
                storage.close()
            QMessageBox.information(self, "PDF export", f"Saved to:\n{path}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", str(exc))

    # ---- search ---------------------------------------------------------
    def _search_changed(self, text: str) -> None:
        needle = (text or "").strip()
        if len(needle) < 2:
            if self._search_override:
                self._search_override = False
                self.stack.setCurrentIndex(self._saved_page)
            return
        if not self._search_override:
            self._search_override = True
            self._saved_page = self.stack.currentIndex()
        self.search_page._storage = self.storage
        self.stack.setCurrentWidget(self.search_page)
        self.search_page.apply_filter(text)

    # ---- auto poll ------------------------------------------------------
    def _toggle_auto_poll(self, checked: bool) -> None:
        if checked:
            self.auto_timer.start(self._auto_interval_ms())
            self.set_status(f"Auto-poll every {self.config['poll']['interval_minutes']} min.")
        else:
            self.auto_timer.stop()
            self.set_status("Auto-poll off.")

    def _auto_interval_ms(self) -> int:
        return max(60, int(self.config["poll"]["interval_minutes"]) * 60 * 1000)

    def _restart_auto_timer(self) -> None:
        if self.auto_timer.isActive():
            self.auto_timer.start(self._auto_interval_ms())


def main() -> int:
    parser = argparse.ArgumentParser(description="ModDB Tracker GUI")
    parser.add_argument("--config", default=CONFIG_FILE, help="path to config.json")
    parser.add_argument("--minimized", action="store_true", help="start hidden in the system tray")
    args = parser.parse_args()

    app = QApplication(sys.argv)
    app.setStyleSheet(QSS)
    transport.patch_moddb()
    window = TrackerWindow(args.config)

    if QSystemTrayIcon.isSystemTrayAvailable():
        app.setQuitOnLastWindowClosed(False)  # closing the window keeps the app in the tray
        if args.minimized or window.config.get("tray", {}).get("start_minimized", False):
            window.hide_to_tray(first_time=True)
            return app.exec()
        window.show()
    else:
        window.show()

    return app.exec()


if __name__ == "__main__":
    sys.exit(main())
